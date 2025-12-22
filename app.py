from flask import Flask, jsonify, request, send_from_directory, send_file
from flask_cors import CORS
import psycopg2
import os
from datetime import datetime, timedelta
import logging
from dotenv import load_dotenv
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import requests
import textwrap

load_dotenv()

app = Flask(__name__, static_folder='frontend')
CORS(app)

TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN', '8281804030:AAEFEYgqigL3bdH4DL0zl1tW71fwwo_8cyU')
ADMIN_TELEGRAM_ID = 174046571  # ← ТВОЙ АЙДИ ЖЕСТКО В КОДЕ

from campaign_calculator import (
    calculate_campaign_price_and_reach,
    STATION_COVERAGE,
    TIME_SLOTS_DATA,
    PRODUCTION_OPTIONS,
    format_number
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def get_db_connection():
    """Создание подключения к PostgreSQL"""
    try:
        conn = psycopg2.connect(os.environ["POSTGRES_URL"])
        return conn
    except Exception as e:
        logger.error(f"❌ Ошибка подключения к БД: {e}")
        return None

def init_db():
    """Инициализация базы данных PostgreSQL"""
    try:
        conn = get_db_connection()
        if not conn:
            return False
            
        cursor = conn.cursor()
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS campaigns (
                id SERIAL PRIMARY KEY,
                user_id BIGINT,
                campaign_number TEXT UNIQUE,
                radio_stations TEXT,
                start_date TEXT,
                end_date TEXT,
                campaign_days INTEGER,
                time_slots TEXT,
                campaign_text TEXT,
                production_option TEXT,
                contact_name TEXT,
                company TEXT,
                phone TEXT,
                email TEXT,
                duration INTEGER,
                base_price INTEGER,
                discount INTEGER,
                final_price INTEGER,
                actual_reach INTEGER,
                status TEXT DEFAULT 'active',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_user_id ON campaigns(user_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_campaign_number ON campaigns(campaign_number)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_created_at ON campaigns(created_at)")
        
        conn.commit()
        cursor.close()
        conn.close()
        logger.info("✅ База данных инициализирована успешно")
        return True
        
    except Exception as e:
        logger.error(f"❌ Ошибка инициализации БД: {e}")
        return False

def send_telegram_to_admin(campaign_number, user_data):
    """ОТПРАВКА УВЕДОМЛЕНИЯ АДМИНУ В TELEGRAM"""
    try:
        stations_text = "\n".join([f"• {radio}" for radio in user_data.get("selected_radios", [])])
        
        final_price = user_data.get('final_price', 0)
        total_reach = user_data.get('total_reach', 0)
        cpc = 0.0
        if total_reach > 0:
            cpc = round(final_price / total_reach, 2)

        notification_text = f"""
🔔 НОВАЯ ЗАЯВКА ИЗ MINI APP #{campaign_number}

👤 КЛИЕНТ:
Имя: {user_data.get('contact_name', 'Не указано')}
Телефон: {user_data.get('phone', 'Не указан')}
Email: {user_data.get('email', 'Не указан')}
Компания: {user_data.get('company', 'Не указана')}

📊 РАДИОСТАНЦИИ:
{stations_text}

📅 ПЕРИОД: {user_data.get('start_date')} - {user_data.get('end_date')} ({user_data.get('campaign_days')} дней)
💰 СТОИМОСТЬ: {format_number(final_price)}₽
👥 ОХВАТ: ~{format_number(total_reach)} чел.
👤 ЦЕНА КОНТАКТА: {cpc}₽
"""
        
        text_url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        text_data = {
            'chat_id': ADMIN_TELEGRAM_ID,
            'text': notification_text,
            'parse_mode': 'HTML'
        }
        response = requests.post(text_url, data=text_data)
        
        if response.status_code != 200:
            logger.error(f"❌ Ошибка отправки текста в Telegram: {response.text}")
            return False
        
        excel_buffer = create_excel_file_from_db(campaign_number)
        if excel_buffer:
            files = {'document': (f'mediaplan_{campaign_number}.xlsx', excel_buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            doc_data = {'chat_id': ADMIN_TELEGRAM_ID}
            doc_url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
            doc_response = requests.post(doc_url, files=files, data=doc_data)
            
            if doc_response.status_code != 200:
                logger.error(f"❌ Ошибка отправки файла в Telegram: {doc_response.text}")
        
        logger.info(f"✅ Уведомление отправлено админу для кампании #{campaign_number}")
        return True
        
    except Exception as e:
        logger.error(f"❌ Ошибка отправки уведомления админу: {e}")
        return False

def send_excel_to_client(campaign_number, user_telegram_id):
    """ОТПРАВКА ТЕКСТА И EXCEL КЛИЕНТУ В TELEGRAM"""
    try:
        # 1. Сначала получаем данные о кампании для текста
        conn = get_db_connection()
        if not conn: return False
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM campaigns WHERE campaign_number = %s", (campaign_number,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()

        if not row: return False

        # Разбираем данные из базы (индексы соответствуют порядку в CREATE TABLE)
        # 3: stations, 4: start, 5: end, 6: days, 17: final_price, 18: reach
        stations_list = row[3].split(',') if row[3] else []
        stations_text = "\n".join([f"• {s}" for s in stations_list])
        start_date = row[4]
        end_date = row[5]
        days = row[6]
        final_price = row[17]
        reach = row[18]
        
        # Считаем цену контакта
        cpc = 0.0
        if reach and reach > 0:
            cpc = round(final_price / reach, 2)

        # 2. Формируем красивый текст для клиента
        client_text = f"""
✅ <b>ВАША ЗАЯВКА #{campaign_number} ПРИНЯТА!</b>

📊 <b>ПАРАМЕТРЫ:</b>
{stations_text}

📅 <b>ПЕРИОД:</b> {start_date} - {end_date} ({days} дней)
💰 <b>СТОИМОСТЬ:</b> {format_number(final_price)}₽
👥 <b>ОХВАТ:</b> ~{format_number(reach)} чел.
👤 <b>ЦЕНА КОНТАКТА:</b> {cpc}₽

📎 <i>Ваш подробный медиаплан во вложении.</i>
📞 <i>Менеджер свяжется с вами в ближайшее время для подтверждения.</i>
"""

        # 3. Отправляем ТЕКСТ клиенту
        requests.post(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage", json={
            'chat_id': user_telegram_id,
            'text': client_text,
            'parse_mode': 'HTML'
        })

        # 4. Отправляем EXCEL клиенту
        excel_buffer = create_excel_file_from_db(campaign_number)
        if excel_buffer and user_telegram_id:
            files = {'document': (f'mediaplan_{campaign_number}.xlsx', excel_buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            doc_data = {'chat_id': user_telegram_id}
            doc_url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
            doc_response = requests.post(doc_url, files=files, data=doc_data)
            
            if doc_response.status_code == 200:
                logger.info(f"✅ Excel отправлен клиенту {user_telegram_id} для кампании #{campaign_number}")
                return True
            else:
                logger.error(f"❌ Ошибка отправки Excel клиенту: {doc_response.text}")
                return False
        return False
        
    except Exception as e:
        logger.error(f"❌ Ошибка отправки Excel клиенту: {e}")
        return False

def create_excel_file_from_db(campaign_number):
    """СОЗДАНИЕ EXCEL МЕДИАПЛАНА"""
    try:
        conn = get_db_connection()
        if not conn:
            return None
            
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM campaigns WHERE campaign_number = %s", (campaign_number,))
        campaign_data = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not campaign_data:
            return None
            
        user_data = {
            "selected_radios": campaign_data[3].split(",") if campaign_data[3] else [],
            "start_date": campaign_data[4],
            "end_date": campaign_data[5],
            "campaign_days": campaign_data[6],
            "selected_time_slots": list(map(int, campaign_data[7].split(","))) if campaign_data[7] else [],
            "campaign_text": campaign_data[8],
            "production_option": campaign_data[9],
            "contact_name": campaign_data[10],
            "company": campaign_data[11],
            "phone": campaign_data[12],
            "email": campaign_data[13],
            "duration": campaign_data[14],
            "base_price": campaign_data[15],
            "discount": campaign_data[16],
            "final_price": campaign_data[17],
            "actual_reach": campaign_data[18]
        }
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Медиаплан {campaign_number}"
        
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=12)
        
        ws.merge_cells("A1:B1")
        ws["A1"] = f"МЕДИАПЛАН КАМПАНИИ #{campaign_number}"
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill
        ws["A1"].alignment = Alignment(horizontal="center")
        
        ws.merge_cells("A2:B2")
        ws["A2"] = "РАДИО ТЮМЕНСКОЙ ОБЛАСТИ"
        ws["A2"].font = title_font
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.append([])
        
        # 📊 ПАРАМЕТРЫ КАМПАНИИ
        current_row = 6
        ws.merge_cells(f"A{current_row}:B{current_row}")
        ws[f"A{current_row}"] = "📊 ПАРАМЕТРЫ КАМПАНИИ:"
        ws[f"A{current_row}"].font = title_font
        current_row += 1
        
        ws[f"A{current_row}"] = "• Радиостанции: " + ", ".join(user_data["selected_radios"])
        current_row += 1
        ws[f"A{current_row}"] = f"• Период: {user_data['start_date']} - {user_data['end_date']} ({user_data['campaign_days']} дней)"
        current_row += 1
        spots_per_day = len(user_data["selected_time_slots"]) * len(user_data["selected_radios"])
        ws[f"A{current_row}"] = f"• Выходов в день: {spots_per_day}"
        current_row += 1
        ws[f"A{current_row}"] = f"• Всего выходов за период: {spots_per_day * user_data['campaign_days']}"
        current_row += 1
        ws[f"A{current_row}"] = f"• Хронометраж ролика: {user_data['duration']} сек"
        current_row += 1
        
        if user_data["campaign_text"] and user_data["campaign_text"].strip():
            ws[f"A{current_row}"] = "• Текст ролика:"
            current_row += 1
            text_lines = textwrap.wrap(user_data["campaign_text"].strip(), width=70)
            for line in text_lines:
                ws[f"A{current_row}"] = f"  {line}"
                current_row += 1
            current_row += 1
        
        production_name = PRODUCTION_OPTIONS.get(user_data["production_option"], {}).get("name", "Не выбрано")
        ws[f"A{current_row}"] = f"• Производство: {production_name}"
        current_row += 2
        
        # РАСЧЕТ ОХВАТА И ЦЕНЫ
        calculation_data = {
            "selected_radios": user_data["selected_radios"],
            "selected_time_slots": user_data["selected_time_slots"],
            "campaign_days": user_data["campaign_days"],
            "duration": user_data["duration"]
        }
        base_price, discount, final_price, total_reach, daily_coverage, spots_per_day, total_coverage_percent, premium_count, cost_per_contact = calculate_campaign_price_and_reach(calculation_data)
        
        # 🎯 РАСЧЕТНЫЙ ОХВАТ
        ws.merge_cells(f"A{current_row}:B{current_row}")
        ws[f"A{current_row}"] = "🎯 РАСЧЕТНЫЙ ОХВАТ:"
        ws[f"A{current_row}"].font = title_font
        current_row += 1
        ws[f"A{current_row}"] = f"• Общий охват за период: ~{format_number(total_reach)} чел."
        current_row += 2
        
        # 💰 ФИНАНСОВАЯ ИНФОРМАЦИЯ
        ws.merge_cells(f"A{current_row}:B{current_row}")
        ws[f"A{current_row}"] = "💰 ФИНАНСОВАЯ ИНФОРМАЦИЯ:"
        ws[f"A{current_row}"].font = title_font
        current_row += 1
        
        ws[f"A{current_row}"] = "Позиция"
        ws[f"B{current_row}"] = "Сумма (₽)"
        current_row += 1
        
        production_cost = PRODUCTION_OPTIONS.get(user_data["production_option"], {}).get('price', 0)
        air_cost = user_data["base_price"] - production_cost
        
        ws[f"A{current_row}"] = "Эфирное время"
        ws[f"B{current_row}"] = air_cost
        current_row += 1
        
        if user_data["production_option"]:
            ws[f"A{current_row}"] = "Производство ролика"
            ws[f"B{current_row}"] = production_cost
            current_row += 1
            
        current_row += 1
        ws[f"A{current_row}"] = "Базовая стоимость"
        ws[f"B{current_row}"] = user_data["base_price"]
        current_row += 1
        
        # СТОИМОСТЬ КОНТАКТА
        ws[f"A{current_row}"] = "Стоимость 1 контакта"
        ws[f"B{current_row}"] = cost_per_contact
        current_row += 1
        
        current_row += 1
        ws[f"A{current_row}"] = "ИТОГО"
        ws[f"B{current_row}"] = user_data["final_price"]
        ws[f"A{current_row}"].font = Font(bold=True)
        ws[f"B{current_row}"].font = Font(bold=True)
        current_row += 3
        
        # 👤 ВАШИ КОНТАКТЫ
        ws.merge_cells(f"A{current_row}:B{current_row}")
        ws[f"A{current_row}"] = "👤 ВАШИ КОНТАКТЫ:"
        ws[f"A{current_row}"].font = title_font
        current_row += 1
        ws[f"A{current_row}"] = f"• Имя: {user_data['contact_name']}"
        current_row += 1
        ws[f"A{current_row}"] = f"• Телефон: {user_data['phone']}"
        current_row += 1
        ws[f"A{current_row}"] = f"• Email: {user_data['email']}"
        current_row += 1
        ws[f"A{current_row}"] = f"• Компания: {user_data['company']}"
        current_row += 2
        
        # 📅 ДАТА
        ws.merge_cells(f"A{current_row}:B{current_row}")
        ws[f"A{current_row}"] = f"📅 Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws[f"A{current_row}"].font = Font(size=9, italic=True)
        
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        logger.error(f"❌ Ошибка при создании Excel: {e}")
        return None

@app.route('/')
def serve_frontend():
    return send_from_directory('frontend', 'index.html')

@app.route('/<path:filename>')
def serve_static(filename):
    return send_from_directory('frontend', filename)

@app.route('/api/health')
def health_check():
    return jsonify({
        "status": "healthy", 
        "database": "connected" if init_db() else "error",
        "timestamp": datetime.now().isoformat()
    })

@app.route('/api/calculate', methods=['POST'])
def calculate_campaign():
    """Расчет стоимости кампании"""
    try:
        data = request.json
        user_data = {
            "selected_radios": data.get('selected_radios', []),
            "start_date": data.get('start_date'),
            "end_date": data.get('end_date'),
            "campaign_days": data.get('campaign_days', 30),
            "selected_time_slots": data.get('selected_time_slots', []),
            "duration": data.get('duration', 20),
            "production_option": data.get('production_option'),
            "production_cost": PRODUCTION_OPTIONS.get(data.get('production_option'), {}).get('price', 0)
        }
        
        base_price, discount, final_price, total_reach, daily_coverage, spots_per_day, total_coverage_percent, premium_count, cost_per_contact = calculate_campaign_price_and_reach(user_data)
        
        return jsonify({
            "success": True,
            "calculation": {
                "base_price": base_price,
                "discount": discount,
                "final_price": final_price,
                "total_reach": total_reach,
                "daily_coverage": daily_coverage,
                "spots_per_day": spots_per_day,
                "total_coverage_percent": total_coverage_percent,
                "premium_count": premium_count,
                "cost_per_contact": cost_per_contact
            }
        })
        
    except Exception as e:
        logger.error(f"Ошибка расчета стоимости: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/create-campaign', methods=['POST'])
def create_campaign():
    """СОЗДАНИЕ НОВОЙ КАМПАНИИ С ЛИМИТОМ 2 В ДЕНЬ"""
    try:
        if not init_db():
            return jsonify({"success": False, "error": "Ошибка инициализации базы данных"}), 500
            
        data = request.json
        user_id = data.get('user_id', 0)
        user_telegram_id = data.get('user_telegram_id')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if user_id == 174046571:
            pass # АДМИН БЕЗ ЛИМИТА
        else:
            cursor.execute("""
                SELECT COUNT(*) FROM campaigns 
                WHERE user_id = %s AND created_at >= NOW() - INTERVAL '1 day'
            """, (user_id,))
            count = cursor.fetchone()[0]
            
            if count >= 2:
                cursor.close()
                conn.close()
                return jsonify({
                    "success": False, 
                    "error": "Превышен лимит в 2 заявки в день. Попробуйте завтра."
                }), 400
        
        calculation_data = {
            "selected_radios": data.get('selected_radios', []),
            "selected_time_slots": data.get('selected_time_slots', []),
            "campaign_days": data.get('campaign_days', 30),
            "duration": data.get('duration', 20),
            "production_option": data.get('production_option'),
            "production_cost": PRODUCTION_OPTIONS.get(data.get('production_option'), {}).get('price', 0)
        }
        
        base_price, discount, final_price, total_reach, daily_coverage, spots_per_day, total_coverage_percent, premium_count, cost_per_contact = calculate_campaign_price_and_reach(calculation_data)
        
        campaign_number = f"R-{datetime.now().strftime('%H%M%S')}"
        
        cursor.execute("""
            INSERT INTO campaigns 
            (user_id, campaign_number, radio_stations, start_date, end_date, campaign_days,
             time_slots, campaign_text, production_option, contact_name,
             company, phone, email, duration, base_price, discount, final_price, actual_reach)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            user_id,
            campaign_number,
            ",".join(data.get("selected_radios", [])),
            data.get("start_date"),
            data.get("end_date"),
            data.get("campaign_days"),
            ",".join(map(str, data.get("selected_time_slots", []))),
            data.get("campaign_text", ""),
            data.get("production_option", ""),
            data.get("contact_name", ""),
            data.get("company", ""),
            data.get("phone", ""),
            data.get("email", ""),
            data.get("duration", 20),
            base_price,
            discount,
            final_price,
            total_reach
        ))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        notification_data = data.copy()
        notification_data['final_price'] = final_price
        notification_data['total_reach'] = total_reach
        send_telegram_to_admin(campaign_number, notification_data)
        
        if user_telegram_id:
            send_excel_to_client(campaign_number, user_telegram_id)
        
        return jsonify({
            "success": True,
            "campaign_number": campaign_number,
            "calculation": {
                "base_price": base_price,
                "discount": discount,
                "final_price": final_price,
                "total_reach": total_reach,
                "cost_per_contact": cost_per_contact
            }
        })
        
    except Exception as e:
        logger.error(f"❌ Ошибка создания кампании: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/user-campaigns/<int:user_id>')
def get_user_campaigns(user_id):
    try:
        if not init_db():
            return jsonify({"success": False, "error": "Ошибка инициализации базы данных"}), 500
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT campaign_number, start_date, end_date, final_price, actual_reach, status, created_at
            FROM campaigns WHERE user_id = %s ORDER BY created_at DESC
        """, (user_id,))
        campaigns = []
        for row in cursor.fetchall():
            campaigns.append({
                "campaign_number": row[0],
                "start_date": row[1],
                "end_date": row[2],
                "final_price": row[3],
                "actual_reach": row[4],
                "status": row[5],
                "created_at": row[6]
            })
        cursor.close()
        conn.close()
        return jsonify({"success": True, "campaigns": campaigns})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/send-excel/<campaign_number>', methods=['POST'])
def send_excel_to_user(campaign_number):
    try:
        data = request.json
        user_telegram_id = data.get('user_telegram_id')
        if not user_telegram_id:
            return jsonify({"success": False, "error": "Не указан Telegram ID"}), 400
        success = send_excel_to_client(campaign_number, user_telegram_id)
        if success:
            return jsonify({"success": True, "message": "Excel отправлен"})
        else:
            return jsonify({"success": False, "error": "Ошибка отправки"}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/delete-campaign/<campaign_number>', methods=['DELETE'])
def delete_campaign(campaign_number):
    try:
        if not init_db():
            return jsonify({"success": False, "error": "DB Error"}), 500
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT status, start_date FROM campaigns WHERE campaign_number = %s", (campaign_number,))
        campaign = cursor.fetchone()
        if not campaign:
            cursor.close(); conn.close()
            return jsonify({"success": False, "error": "Кампания не найдена"}), 404
        status, start_date = campaign
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d') if start_date else datetime.now()
        if status != 'active' or start_date_obj <= datetime.now():
            cursor.close(); conn.close()
            return jsonify({"success": False, "error": "Нельзя удалить"}), 400
        cursor.execute("DELETE FROM campaigns WHERE campaign_number = %s", (campaign_number,))
        conn.commit()
        cursor.close(); conn.close()
        return jsonify({"success": True, "message": "Удалено"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/confirmation/<campaign_number>')
def get_campaign_confirmation(campaign_number):
    try:
        if not init_db(): return jsonify({"success": False, "error": "DB Error"}), 500
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT campaign_number, final_price, actual_reach, contact_name, phone, email, company, created_at, start_date, end_date FROM campaigns WHERE campaign_number = %s", (campaign_number,))
        campaign = cursor.fetchone()
        cursor.close(); conn.close()
        if not campaign: return jsonify({"success": False, "error": "Not found"}), 404
        
        final_price = campaign[1]
        actual_reach = campaign[2]
        cpc = round(final_price / actual_reach, 2) if actual_reach > 0 else 0
        
        return jsonify({
            "success": True,
            "campaign": {
                "campaign_number": campaign[0],
                "final_price": final_price,
                "actual_reach": actual_reach,
                "contact_name": campaign[3],
                "phone": campaign[4],
                "email": campaign[5],
                "company": campaign[6],
                "created_at": campaign[7],
                "start_date": campaign[8],
                "end_date": campaign[9],
                "cost_per_contact": cpc
            }
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)

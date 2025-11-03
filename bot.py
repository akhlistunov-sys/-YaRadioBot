import os
import logging
import sqlite3
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes, ConversationHandler
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Состояния разговора
MAIN_MENU, RADIO_SELECTION, CAMPAIGN_DATES, TIME_SLOTS, BRANDED_SECTIONS, CAMPAIGN_CREATOR, PRODUCTION_OPTION, CONTACT_INFO, CONFIRMATION, FINAL_ACTIONS = range(10)

# Токен бота
TOKEN = "8281804030:AAEFEYgqigL3bdH4DL0zl1tW71fwwo_8cyU"

# Ваш Telegram ID для уведомлений
ADMIN_TELEGRAM_ID = 174046571

# Цены и параметры
BASE_PRICE_PER_SECOND = 2.0  # 2₽ за секунду
MIN_PRODUCTION_COST = 2000
MIN_BUDGET = 7000

# Обновленные данные слотов с разным охватом
TIME_SLOTS_DATA = [
    {"time": "06:00-07:00", "label": "Подъем, сборы", "premium": True, "coverage_percent": 6},
    {"time": "07:00-08:00", "label": "Утренние поездки", "premium": True, "coverage_percent": 10},
    {"time": "08:00-09:00", "label": "Пик трафика", "premium": True, "coverage_percent": 12},
    {"time": "09:00-10:00", "label": "Начало работы", "premium": True, "coverage_percent": 8},
    {"time": "10:00-11:00", "label": "Рабочий процесс", "premium": True, "coverage_percent": 7},
    {"time": "11:00-12:00", "label": "Предобеденное время", "premium": True, "coverage_percent": 6},
    {"time": "12:00-13:00", "label": "Обеденный перерыв", "premium": True, "coverage_percent": 5},
    {"time": "13:00-14:00", "label": "После обеда", "premium": True, "coverage_percent": 5},
    {"time": "14:00-15:00", "label": "Вторая половина дня", "premium": True, "coverage_percent": 5},
    {"time": "15:00-16:00", "label": "Рабочий финиш", "premium": True, "coverage_percent": 6},
    {"time": "16:00-17:00", "label": "Конец рабочего дня", "premium": True, "coverage_percent": 7},
    {"time": "17:00-18:00", "label": "Вечерние поездки", "premium": True, "coverage_percent": 10},
    {"time": "18:00-19:00", "label": "Пик трафика", "premium": True, "coverage_percent": 8},
    {"time": "19:00-20:00", "label": "Домашний вечер", "premium": True, "coverage_percent": 4},
    {"time": "20:00-21:00", "label": "Вечерний отдых", "premium": True, "coverage_percent": 4}
]

# Обновленные данные охвата (усредненные)
STATION_COVERAGE = {
    "LOVE RADIO": 540,
    "АВТОРАДИО": 3250,
    "РАДИО ДАЧА": 3250,
    "РАДИО ШАНСОН": 2900,
    "РЕТРО FM": 3600,
    "ЮМОР FM": 1260
}

BRANDED_SECTION_PRICES = {
    "auto": 1.2,
    "realty": 1.15,
    "medical": 1.25,
    "custom": 1.3
}

PRODUCTION_OPTIONS = {
    "standard": {"price": 2000, "name": "СТАНДАРТНЫЙ РОЛИК", "desc": "Профессиональная озвучка, музыкальное оформление, срок: 2-3 дня"},
    "premium": {"price": 5000, "name": "ПРЕМИУМ РОЛИК", "desc": "Озвучка 2-мя голосами, индивидуальная музыка, срочное производство 1 день"}
}

# Инициализация базы данных
def init_db():
    try:
        conn = sqlite3.connect("campaigns.db")
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS campaigns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                campaign_number TEXT,
                radio_stations TEXT,
                start_date TEXT,
                end_date TEXT,
                campaign_days INTEGER,
                time_slots TEXT,
                branded_section TEXT,
                campaign_text TEXT,
                production_option TEXT,
                contact_name TEXT,
                company TEXT,
                phone TEXT,
                email TEXT,
                duration INTEGER,
                base_price INTEGER,
                discount INTEGER,
                final_price INTEGER,
                actual_reach INTEGER,
                status TEXT DEFAULT "active",
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Таблица для ограничения частоты запросов
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS rate_limits (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                action_type TEXT,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        conn.commit()
        conn.close()
        logger.info("База данных инициализирована успешно")
        return True
    except Exception as e:
        logger.error(f"Ошибка инициализации БД: {e}")
        return False

def validate_phone(phone: str) -> bool:
    """Упрощенная валидация телефона"""
    if not phone:
        return False
    return True

def validate_date(date_text: str) -> bool:
    """Проверка валидности даты"""
    try:
        date = datetime.strptime(date_text, "%d.%m.%Y")
        if date < datetime.now().replace(hour=0, minute=0, second=0, microsecond=0):
            return False
        if date > datetime.now() + timedelta(days=365):
            return False
        return True
    except ValueError:
        return False

def format_number(num):
    return f"{num:,}".replace(",", " ")

def check_rate_limit(user_id: int) -> bool:
    """Проверка ограничения в 5 заявок в день"""
    try:
        conn = sqlite3.connect("campaigns.db")
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT COUNT(*) FROM campaigns 
            WHERE user_id = ? AND created_at >= datetime("now", "-1 day")
        """, (user_id,))
        
        count = cursor.fetchone()[0]
        conn.close()
        
        return count < 5
    except Exception as e:
        logger.error(f"Ошибка проверки лимитa: {e}")
        return True

def calculate_campaign_price_and_reach(user_data):
    """ОБНОВЛЕННАЯ ФУНКЦИЯ РАСЧЕТА С РАЗНЫМ ОХВАТОМ СЛОТОВ"""
    try:
        base_duration = user_data.get("duration", 20)
        campaign_days = user_data.get("campaign_days", 30)
        selected_radios = user_data.get("selected_radios", [])
        selected_time_slots = user_data.get("selected_time_slots", [])
        
        if not selected_radios or not selected_time_slots:
            return 0, 0, MIN_BUDGET, 0, 0, 0, 0
            
        num_stations = len(selected_radios)
        spots_per_day = len(selected_time_slots) * num_stations
        
        cost_per_spot = base_duration * BASE_PRICE_PER_SECOND
        base_air_cost = cost_per_spot * spots_per_day * campaign_days
        
        time_multiplier = 1.0
        for slot_index in selected_time_slots:
            if 0 <= slot_index < len(TIME_SLOTS_DATA):
                slot = TIME_SLOTS_DATA[slot_index]
                if slot["premium"]:
                    time_multiplier = max(time_multiplier, 1.1)
        
        branded_multiplier = 1.0
        branded_section = user_data.get("branded_section")
        if branded_section in BRANDED_SECTION_PRICES:
            branded_multiplier = BRANDED_SECTION_PRICES[branded_section]
        
        production_cost = user_data.get("production_cost", 0)
        air_cost = int(base_air_cost * time_multiplier * branded_multiplier)
        base_price = air_cost + production_cost
        
        discount = int(base_price * 0.5)
        discounted_price = base_price - discount
        final_price = max(discounted_price, MIN_BUDGET)
        
        # ОБНОВЛЕННЫЙ РАСЧЕТ ОХВАТА С РАЗНЫМИ % СЛОТОВ
        total_listeners = sum(STATION_COVERAGE.get(radio, 0) for radio in selected_radios)
        
        # Сумма % охвата выбранных слотов
        total_coverage_percent = 0
        for slot_index in selected_time_slots:
            if 0 <= slot_index < len(TIME_SLOTS_DATA):
                slot = TIME_SLOTS_DATA[slot_index]
                total_coverage_percent += slot["coverage_percent"]
        
        # Уникальный охват с учетом пересечения аудитории (0.7)
        unique_daily_coverage = int(total_listeners * 0.7 * (total_coverage_percent / 100))
        total_reach = int(unique_daily_coverage * campaign_days)
        
        return base_price, discount, final_price, total_reach, unique_daily_coverage, spots_per_day, total_coverage_percent
        
    except Exception as e:
        logger.error(f"Ошибка расчета стоимости: {e}")
        return 0, 0, MIN_BUDGET, 0, 0, 0, 0

def get_branded_section_name(section):
    names = {
        "auto": "Авторубрики (+20%)",
        "realty": "Недвижимость (+15%)",
        "medical": "Медицинские рубрики (+25%)",
        "custom": "Индивидуальная рубрика (+30%)"
    }
    return names.get(section, "Не выбрана")

def get_time_slots_text(selected_slots):
    """Получить текстовое представление выбранных слотов"""
    slots_text = ""
    for slot_index in selected_slots:
        if 0 <= slot_index < len(TIME_SLOTS_DATA):
            slot = TIME_SLOTS_DATA[slot_index]
            premium_emoji = "🚀" if slot["premium"] else "📊"
            slots_text += f"• {slot['time']} - {slot['label']} {premium_emoji}\n"
    return slots_text

def get_time_slots_detailed_text(selected_slots):
    """Получить детальное представление слотов с охватом"""
    slots_text = ""
    total_coverage = 0
    
    for slot_index in selected_slots:
        if 0 <= slot_index < len(TIME_SLOTS_DATA):
            slot = TIME_SLOTS_DATA[slot_index]
            premium_emoji = "🚀" if slot["premium"] else "📊"
            coverage_percent = slot["coverage_percent"]
            total_coverage += coverage_percent
            slots_text += f"• {slot['time']} - {slot['label']}: {coverage_percent}% {premium_emoji}\n"
    
    return slots_text, total_coverage

def create_excel_file_from_db(campaign_number):
    try:
        logger.info(f"🔍 Начинаем создание Excel для кампании #{campaign_number}")
        
        conn = sqlite3.connect("campaigns.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM campaigns WHERE campaign_number = ?", (campaign_number,))
        campaign_data = cursor.fetchone()
        conn.close()
        
        if not campaign_data:
            logger.error(f"❌ Кампания #{campaign_number} не найдена в БД")
            return None
            
        logger.info(f"✅ Кампания #{campaign_number} найдена в БД")
        
        user_data = {
            "selected_radios": campaign_data[3].split(","),
            "start_date": campaign_data[4],
            "end_date": campaign_data[5],
            "campaign_days": campaign_data[6],
            "selected_time_slots": list(map(int, campaign_data[7].split(","))) if campaign_data[7] else [],
            "branded_section": campaign_data[8],
            "campaign_text": campaign_data[9],
            "production_option": campaign_data[10],
            "contact_name": campaign_data[11],
            "company": campaign_data[12],
            "phone": campaign_data[13],
            "email": campaign_data[14],
            "duration": campaign_data[15],
            "production_cost": PRODUCTION_OPTIONS.get(campaign_data[10], {}).get("price", 0)
        }
        
        logger.info(f"📊 Данные пользователя подготовлены: {len(user_data.get('selected_radios', []))} радиостанций")
        
        base_price, discount, final_price, total_reach, daily_coverage, spots_per_day, total_coverage_percent = calculate_campaign_price_and_reach(user_data)
        logger.info(f"💰 Расчет стоимости: база={base_price}, скидка={discount}, итого={final_price}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Медиаплан {campaign_number}"
        
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=12)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                       top=Side(style="thin"), bottom=Side(style="thin"))
        
        ws.merge_cells("A1:F1")
        ws["A1"] = f"МЕДИАПЛАН КАМПАНИИ #{campaign_number}"
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill
        ws["A1"].alignment = Alignment(horizontal="center")
        
        ws.merge_cells("A2:F2")
        ws["A2"] = "РАДИО ТЮМЕНСКОЙ ОБЛАСТИ"
        ws["A2"].font = Font(bold=True, size=12, color="366092")
        ws["A2"].alignment = Alignment(horizontal="center")
        
        ws.merge_cells("A4:F4")
        ws["A4"] = "✅ Ваша заявка принята! Спасибо за доверие!"
        ws["A4"].font = Font(bold=True, size=11)
        
        ws["A6"] = "📊 ПАРАМЕТРЫ КАМПАНИИ:"
        ws["A6"].font = title_font
        
        # Детальная информация о выбранных слотах
        slots_text, total_coverage = get_time_slots_detailed_text(user_data.get("selected_time_slots", []))
        
        params = [
            f"Радиостанции: {', '.join(user_data.get('selected_radios', []))}",
            f"Период: {user_data.get('start_date')} - {user_data.get('end_date')} ({user_data.get('campaign_days')} дней)",
            f"Выходов в день: {spots_per_day}",
            f"Всего выходов за период: {spots_per_day * user_data.get('campaign_days', 30)}",
            f"Хронометраж ролика: {user_data.get('duration', 20)} сек",
            f"Брендированная рубрика: {get_branded_section_name(user_data.get('branded_section'))}",
            f"Производство: {PRODUCTION_OPTIONS.get(user_data.get('production_option', 'ready'), {}).get('name', 'Не выбрано')}",
            f"Суммарный охват выбранных слотов: {total_coverage}%"
        ]
        
        for i, param in enumerate(params, 7):
            ws[f"A{i}"] = f"• {param}"
        
        ws["A16"] = "📻 ВЫБРАННЫЕ РАДИОСТАНЦИИ:"
        ws["A16"].font = title_font
        
        row = 17
        total_listeners = 0
        for radio in user_data.get("selected_radios", []):
            listeners = STATION_COVERAGE.get(radio, 0)
            total_listeners += listeners
            ws[f"A{row}"] = f"• {radio}: ~{format_number(listeners)} слушателей"
            row += 1
        
        ws[f"A{row}"] = f"• ИТОГО: ~{format_number(total_listeners)} слушателей"
        ws[f"A{row}"].font = Font(bold=True)
        
        row += 2
        ws[f"A{row}"] = "🕒 ВЫБРАННЫЕ ВРЕМЕННЫЕ СЛОТЫ:"
        ws[f"A{row}"].font = title_font
        
        row += 1
        for slot_index in user_data.get("selected_time_slots", []):
            if 0 <= slot_index < len(TIME_SLOTS_DATA):
                slot = TIME_SLOTS_DATA[slot_index]
                ws[f"A{row}"] = f"• {slot['time']} - {slot['label']}: {slot['coverage_percent']}%"
                row += 1
        
        ws[f"A{row}"] = f"• Суммарный охват слотов: {total_coverage}%"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        
        row += 1
        ws[f"A{row}"] = "🎯 РАСЧЕТНЫЙ ОХВАТ:"
        ws[f"A{row}"].font = title_font
        
        row += 1
        ws[f"A{row}"] = f"• Выходов в день: {spots_per_day}"
        row += 1
        ws[f"A{row}"] = f"• Уникальных слушателей в день: ~{format_number(daily_coverage)} чел."
        row += 1
        ws[f"A{row}"] = f"• Общий охват за период: ~{format_number(total_reach)} чел."
        
        row += 2
        ws[f"A{row}"] = "💰 ФИНАНСОВАЯ ИНФОРМАЦИЯ:"
        ws[f"A{row}"].font = title_font
        
        financial_data = [
            ["Позиция", "Сумма (₽)"],
            ["Эфирное время", base_price - user_data.get("production_cost", 0)],
            ["Производство ролика", user_data.get("production_cost", 0)],
            ["", ""],
            ["Базовая стоимость", base_price],
            ["Скидка 50%", -discount],
            ["", ""],
            ["ИТОГО", final_price]
        ]
        
        for i, (item, value) in enumerate(financial_data, row + 1):
            ws[f"A{i}"] = item
            if isinstance(value, int):
                ws[f"B{i}"] = value
                if item == "ИТОГО":
                    ws[f"B{i}"].font = Font(bold=True, color="FF0000")
                elif item == "Скидка 50%":
                    ws[f"B{i}"].font = Font(color="00FF00")
            else:
                ws[f"B{i}"] = value
        
        row = i + 3
        ws[f"A{row}"] = "👤 ВАШИ КОНТАКТЫ:"
        ws[f"A{row}"].font = title_font
        
        contacts = [
            f"Имя: {user_data.get('contact_name', 'Не указано')}",
            f"Телефон: {user_data.get('phone', 'Не указан')}",
            f"Email: {user_data.get('email', 'Не указан')}",
            f"Компания: {user_data.get('company', 'Не указана')}"
        ]
        
        for i, contact in enumerate(contacts, row + 1):
            ws[f"A{i}"] = f"• {contact}"
        
        row = i + 2
        ws[f"A{row}"] = "📞 НАШИ КОНТАКТЫ:"
        ws[f"A{row}"].font = title_font
        ws[f"A{row + 1}"] = "• Email: a.khlistunov@gmail.com"
        ws[f"A{row + 2}"] = "• Telegram: t.me/AlexeyKhlistunov"
        
        row = row + 4
        ws[f"A{row}"] = "🎯 СТАРТ КАМПАНИИ:"
        ws[f"A{row}"].font = title_font
        ws[f"A{row + 1}"] = "В течение 3 рабочих дней после подтверждения"
        
        row = row + 3
        ws[f"A{row}"] = f"📅 Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 15
        
        table_start = row - len(financial_data) - 1
        table_end = table_start + len(financial_data) - 1
        for row_num in range(table_start, table_end + 1):
            for col in ["A", "B"]:
                ws[f"{col}{row_num}"].border = border
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"✅ Excel файл успешно создан для кампании #{campaign_number}, размер: {len(buffer.getvalue())} байт")
        return buffer
        
    except Exception as e:
        logger.error(f"❌ Ошибка при создании Excel: {e}")
        return None

async def send_admin_notification(context, user_data, campaign_number):
    try:
        excel_buffer = create_excel_file_from_db(campaign_number)
        if excel_buffer:
            await context.bot.send_document(
                chat_id=ADMIN_TELEGRAM_ID,
                document=excel_buffer,
                filename=f"mediaplan_{campaign_number}.xlsx",
                caption=f"📊 Медиаплан кампании #{campaign_number}"
            )
            logger.info(f"✅ Excel автоматически отправлен админу для кампании #{campaign_number}")
        
        base_price, discount, final_price, total_reach, daily_coverage, spots_per_day, total_coverage_percent = calculate_campaign_price_and_reach(user_data)
        
        stations_text = ""
        for radio in user_data.get("selected_radios", []):
            listeners = STATION_COVERAGE.get(radio, 0)
            stations_text += f"• {radio}: ~{format_number(listeners)} слушателей\n"
        
        slots_text = get_time_slots_text(user_data.get("selected_time_slots", []))
        
        notification_text = f"""
🔔 НОВАЯ ЗАЯВКА #{campaign_number}

👤 КЛИЕНТ:
Имя: {user_data.get('contact_name', 'Не указано')}
Телефон: {user_data.get('phone', 'Не указан')}
Email: {user_data.get('email', 'Не указан')}
Компания: {user_data.get('company', 'Не указана')}

📊 РАДИОСТАНЦИИ:
{stations_text}
📅 ПЕРИОД: {user_data.get('start_date')} - {user_data.get('end_date')} ({user_data.get('campaign_days')} дней)
🕒 ВЫБРАНО СЛОТОВ: {len(user_data.get('selected_time_slots', []))}
{slots_text}
🎙️ РУБРИКА: {get_branded_section_name(user_data.get('branded_section'))}
⏱️ РОЛИК: {PRODUCTION_OPTIONS.get(user_data.get('production_option', 'ready'), {}).get('name', 'Не выбрано')}
📏 ХРОНОМЕТРАЖ: {user_data.get('duration', 20)} сек

💰 СТОИМОСТЬ:
Базовая: {format_number(base_price)}₽
Скидка 50%: -{format_number(discount)}₽
Итоговая: {format_number(final_price)}₽

🎯 РАСЧЕТНЫЙ ОХВАТ:
• Выходов в день: {spots_per_day}
• Всего выходов: {spots_per_day * user_data.get('campaign_days', 30)}
• Ежедневно: ~{format_number(daily_coverage)} чел.
• За период: ~{format_number(total_reach)} чел.
"""
        
        keyboard = [
            [
                InlineKeyboardButton(f"📞 {user_data.get('phone', 'Телефон')}", callback_data=f"call_{user_data.get('phone', '')}"),
                InlineKeyboardButton(f"✉️ Написать", callback_data=f"email_{user_data.get('email', '')}")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await context.bot.send_message(
            chat_id=ADMIN_TELEGRAM_ID,
            text=notification_text,
            reply_markup=reply_markup
        )
        logger.info(f"✅ Уведомление админу отправлено для кампании #{campaign_number}")
        return True
        
    except Exception as e:
        logger.error(f"❌ Ошибка отправки админу: {e}")
        return False

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ГЛАВНОЕ МЕНЮ"""
    keyboard = [
        [InlineKeyboardButton("🚀 НАЧАТЬ РАСЧЕТ", callback_data="create_campaign")],
        [InlineKeyboardButton("📊 ВОЗРАСТНАЯ СТРУКТУРА", callback_data="statistics")],
        [InlineKeyboardButton("🏆 О НАС", callback_data="about")],
        [InlineKeyboardButton("📋 ЛИЧНЫЙ КАБИНЕТ", callback_data="personal_cabinet")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    caption = (
        "🎙️ РАДИО ТЮМЕНСКОЙ ОБЛАСТИ\n"
        "📍 Ялуторовск • Заводоуковск\n\n"
        "🤖 **Рассчитайте рекламу за 2 минуты**\n"
        "3 простых шага → готовый медиаплан\n\n"
        "• 6 федеральных радиостанций\n"
        "• Скидка 50% на первую кампанию\n"
        "• Старт через 3 дня\n"
        "• Персональный медиаплан\n\n"
        "🏆 70+ кампаний в 2025 году\n"
        "✅ От 7 000₽"
    )
    
    if update.message:
        await update.message.reply_text(
            caption,
            reply_markup=reply_markup
        )
    else:
        query = update.callback_query
        await query.answer()
        await query.edit_message_text(
            caption,
            reply_markup=reply_markup
        )
    
    return MAIN_MENU

async def about_section(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """РАЗДЕЛ О НАС"""
    query = update.callback_query
    await query.answer()
    
    about_text = """🏆 О НАС

«Мы не продаём секунды эфира —
мы создаём истории, которые
слышит весь регион»

10 лет мы помогаем бизнесу
достигать своей аудитории
через силу радиоволн.

📻 6 федеральных станций
📍 Ялуторовск • Заводоуковск  
🎯 40 000+ слушателей

Наша миссия — делать рекламу,
которую слушают, а не пропускают."""
    
    keyboard = [
        [InlineKeyboardButton("🚀 НАЧАТЬ РАСЧЕТ", callback_data="create_campaign")],
        [InlineKeyboardButton("📞 КОНТАКТЫ", callback_data="contacts_details")],
        [InlineKeyboardButton("🔙 НАЗАД", callback_data="back_to_main")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(about_text, reply_markup=reply_markup)
    return MAIN_MENU

async def radio_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ШАГ 1/7 - ВЫБОР РАДИОСТАНЦИЙ С ОПИСАНИЯМИ"""
    query = update.callback_query
    await query.answer()
    
    selected_radios = context.user_data.get("selected_radios", [])
    total_listeners = sum(STATION_COVERAGE.get(radio, 0) for radio in selected_radios)
    
    keyboard = []
    
    all_selected = len(selected_radios) == 6
    keyboard.append([InlineKeyboardButton(
        "✅ ВЫБРАТЬ ВСЕ 6 РАДИОСТАНЦИЙ" if all_selected else "⚪ ВЫБРАТЬ ВСЕ 6 РАДИОСТАНЦИЙ", 
        callback_data="select_all_radios"
    )])
    
    radio_stations = [
        ("LOVE RADIO", "radio_love", 540),
        ("АВТОРАДИО", "radio_auto", 3250),
        ("РАДИО ДАЧА", "radio_dacha", 3250), 
        ("РАДИО ШАНСОН", "radio_chanson", 2900),
        ("РЕТРО FM", "radio_retro", 3600),
        ("ЮМОР FM", "radio_humor", 1260)
    ]
    
    for name, callback, listeners in radio_stations:
        emoji = "✅" if name in selected_radios else "⚪"
        button_text = f"{emoji} {name} (~{format_number(listeners)} слушателей)"
        keyboard.append([InlineKeyboardButton(button_text, callback_data=callback)])
    
    keyboard.append([InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_main")])
    keyboard.append([InlineKeyboardButton("➡️ ДАЛЕЕ", callback_data="to_campaign_dates")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    # ОБНОВЛЕННЫЙ ТЕКСТ С ОПИСАНИЯМИ РАДИОСТАНЦИЙ
    text = (
        f"● ● ○ ○ ○ ○ ○   [1/7] ВЫБОР РАДИОСТАНЦИЙ\n\n"
        f"⏱️ Этот шаг займет ~30 секунд\n\n"
        f"Выбрано: {len(selected_radios)} станции • ~{format_number(total_listeners)} слушателей в день\n\n"
        f"⚪ LOVE RADIO\n"
        f"~540 слушателей в день\n"
        f"👩 Молодёжь 16-35 лет\n\n"
        f"⚪ АВТОРАДИО\n"
        f"~3,250 слушателей в день\n"
        f"👨 Автомобилисты 25-55 лет\n\n"
        f"⚪ РАДИО ДАЧА\n"
        f"~3,250 слушателей в день\n"
        f"👨👩 Семья 35-60 лет\n\n"
        f"⚪ РАДИО ШАНСОН\n"
        f"~2,900 слушателей в день\n"
        f"👨 Мужчины 30-60+ лет\n\n"
        f"⚪ РЕТРО FM\n"
        f"~3,600 слушателей в день\n"
        f"👴👵 Взрослые 35-65 лет\n\n"
        f"⚪ ЮМОР FM\n"
        f"~1,260 слушателей в день\n"
        f"👦👧 Молодежь 12-19 и взрослые 25-45 лет\n\n"
        f"✅ Готово! Следующий шаг: выбор дат (15 сек)"
    )
    
    await query.edit_message_text(text, reply_markup=reply_markup)
    return RADIO_SELECTION

async def handle_radio_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ОБРАБОТЧИК ВЫБОРА РАДИОСТАНЦИЙ"""
    query = update.callback_query
    await query.answer()
    
    if query.data == "back_to_main":
        return await start(update, context)
    
    elif query.data == "select_all_radios":
        all_radios = ["LOVE RADIO", "АВТОРАДИО", "РАДИО ДАЧА", "РАДИО ШАНСОН", "РЕТРО FM", "ЮМОР FM"]
        context.user_data["selected_radios"] = all_radios
        return await radio_selection(update, context)
    
    radio_data = {
        "radio_love": "LOVE RADIO",
        "radio_auto": "АВТОРАДИО", 
        "radio_dacha": "РАДИО ДАЧА",
        "radio_chanson": "РАДИО ШАНСОН",
        "radio_retro": "РЕТРО FM",
        "radio_humor": "ЮМОР FM"
    }
    
    if query.data in radio_data:
        radio_name = radio_data[query.data]
        selected_radios = context.user_data.get("selected_radios", [])
        
        if radio_name in selected_radios:
            selected_radios.remove(radio_name)
        else:
            selected_radios.append(radio_name)
        
        context.user_data["selected_radios"] = selected_radios
        return await radio_selection(update, context)
    
    elif query.data == "to_campaign_dates":
        if not context.user_data.get("selected_radios"):
            await query.answer("❌ Выберите хотя бы одну радиостанцию!", show_alert=True)
            return RADIO_SELECTION
        
        # Автоматический переход к вводу дат
        keyboard = [[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_radio")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        selected_radios = context.user_data.get("selected_radios", [])
        stations_info = "📻 ВЫБРАНЫ СТАНЦИИ:\n"
        for radio in selected_radios:
            listeners = STATION_COVERAGE.get(radio, 0)
            stations_info += f"• {radio}: ~{format_number(listeners)} слушателей в день\n"
        
        text = (
            f"● ● ● ○ ○ ○ ○   [2/7] ВЫБОР ДАТ КАМПАНИИ\n\n"
            f"⏱️ Этот шаг займет ~15 секунд\n\n"
            f"{stations_info}\n"
            f"🗓️ Период не выбран\n\n"
            f"─────────────────\n"
            f"✅ Период: 0 дней\n"
            f"⚠️ Минимальный период: 15 дней\n\n"
            f"📅 Введите дату начала кампании в формате ДД.ММ.ГГГГ:\n\n"
            f"Пример: 15.01.2025\n\n"
            f"Отправьте дату сообщением:"
        )
        
        await query.edit_message_text(text, reply_markup=reply_markup)
        return "WAITING_START_DATE"
    
    return RADIO_SELECTION

async def process_start_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ОБРАБОТКА ДАТЫ НАЧАЛА"""
    try:
        date_text = update.message.text.strip()
        
        if not validate_date(date_text):
            await update.message.reply_text(
                "❌ Неверная дата. Проверьте:\n"
                "• Формат ДД.ММ.ГГГГ\n"
                "• Дата не в прошлом\n"
                "• Дата не более чем на 1 год вперед\n\n"
                "Введите корректную дату:",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_radio")]])
            )
            return "WAITING_START_DATE"
        
        context.user_data["start_date"] = date_text
        
        keyboard = [[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_radio")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            "📅 Введите дату окончания кампании в формате ДД.ММ.ГГГГ:\n\n"
            "Пример: 30.01.2025\n\n"
            "Отправьте дату сообщением:",
            reply_markup=reply_markup
        )
        return "WAITING_END_DATE"
        
    except ValueError:
        await update.message.reply_text(
            "❌ Неверный формат даты. Используйте ДД.ММ.ГГГГ:\n\n"
            "Пример: 15.01.2025",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_radio")]])
        )
        return "WAITING_START_DATE"

async def process_end_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ОБРАБОТКА ДАТЫ ОКОНЧАНИЯ"""
    try:
        date_text = update.message.text.strip()
        
        if not validate_date(date_text):
            await update.message.reply_text(
                "❌ Неверная дата. Проверьте:\n"
                "• Формат ДД.ММ.ГГГГ\n"
                "• Дата не в прошлом\n"
                "• Дата не более чем на 1 год вперед\n\n"
                "Введите корректную дату:",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_radio")]])
            )
            return "WAITING_END_DATE"
        
        if not context.user_data.get("start_date"):
            await update.message.reply_text(
                "❌ Сначала выберите дату начала. Введите дату окончания снова:",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_radio")]])
            )
            return "WAITING_END_DATE"
        
        start_date = datetime.strptime(context.user_data["start_date"], "%d.%m.%Y")
        end_date = datetime.strptime(date_text, "%d.%m.%Y")
        
        if end_date <= start_date:
            await update.message.reply_text(
                "❌ Дата окончания должна быть после даты начала. Введите корректную дату:",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_radio")]])
            )
            return "WAITING_END_DATE"
        
        campaign_days = (end_date - start_date).days + 1
        
        if campaign_days < 15:
            await update.message.reply_text(
                "❌ Минимальный период кампании - 15 дней. Введите дату окончания снова:",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_radio")]])
            )
            return "WAITING_END_DATE"
        
        context.user_data["end_date"] = date_text
        context.user_data["campaign_days"] = campaign_days
        
        # Автоматический переход к выбору времени
        return await time_slots_from_message(update, context)
        
    except ValueError:
        await update.message.reply_text(
            "❌ Неверный формат даты. Используйте ДД.ММ.ГГГГ:\n\n"
            "Пример: 30.01.2025",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_radio")]])
        )
        return "WAITING_END_DATE"

async def handle_campaign_dates(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ОБРАБОТЧИК ВЫБОРА ДАТ КАМПАНИИ"""
    query = update.callback_query
    await query.answer()
    
    if query.data == "back_to_radio":
        return await radio_selection(update, context)
    
    elif query.data == "select_period":
        keyboard = [[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_dates")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            "📅 Введите дату начала кампании в формате ДД.ММ.ГГГГ:\n\n"
            "Пример: 15.01.2025\n\n"
            "Отправьте дату сообщением:",
            reply_markup=reply_markup
        )
        return "WAITING_START_DATE"
    
    elif query.data == "to_time_slots":
        if not context.user_data.get("start_date") or not context.user_data.get("end_date"):
            await query.answer("❌ Выберите период кампании!", show_alert=True)
            return CAMPAIGN_DATES
        
        campaign_days = context.user_data.get("campaign_days", 0)
        if campaign_days < 15:
            await query.answer("❌ Минимальный период кампании - 15 дней!", show_alert=True)
            return CAMPAIGN_DATES
            
        return await time_slots(update, context)
    
    elif query.data == "cancel_period":
        return await campaign_dates(update, context)
    
    return CAMPAIGN_DATES

async def campaign_dates(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ШАГ 2/7 - ВЫБОР ДАТ КАМПАНИИ (РЕЗЕРВНЫЙ)"""
    query = update.callback_query
    await query.answer()
    
    keyboard = [[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_radio")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    selected_radios = context.user_data.get("selected_radios", [])
    stations_info = "📻 ВЫБРАНЫ СТАНЦИИ:\n"
    for radio in selected_radios:
        listeners = STATION_COVERAGE.get(radio, 0)
        stations_info += f"• {radio}: ~{format_number(listeners)} слушателей в день\n"
    
    text = (
        f"● ● ● ○ ○ ○ ○   [2/7] ВЫБОР ДАТ КАМПАНИИ\n\n"
        f"⏱️ Этот шаг займет ~15 секунд\n\n"
        f"{stations_info}\n"
        f"🗓️ Период не выбран\n\n"
        f"─────────────────\n"
        f"✅ Период: 0 дней\n"
        f"⚠️ Минимальный период: 15 дней\n\n"
        f"📅 Введите дату начала кампании в формате ДД.ММ.ГГГГ:\n\n"
        f"Пример: 15.01.2025\n\n"
        f"Отправьте дату сообщением:"
    )
    
    await query.edit_message_text(text, reply_markup=reply_markup)
    return "WAITING_START_DATE"

async def time_slots_from_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ПЕРЕХОД К ВЫБОРУ ВРЕМЕНИ ИЗ СООБЩЕНИЯ"""
    selected_slots = context.user_data.get("selected_time_slots", [])
    selected_radios = context.user_data.get("selected_radios", [])
    campaign_days = context.user_data.get("campaign_days", 30)
    
    keyboard = []
    keyboard.append([InlineKeyboardButton("✅ ВЫБРАТЬ ВСЕ СЛОТЫ", callback_data="select_all_slots")])
    
    for i in range(15):
        slot = TIME_SLOTS_DATA[i]
        emoji = "✅" if i in selected_slots else "⚪"
        button_text = f"{emoji} {slot['time']} • {slot['label']}"
        keyboard.append([InlineKeyboardButton(button_text, callback_data=f"time_{i}")])
    
    keyboard.append([InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_dates")])
    keyboard.append([InlineKeyboardButton("➡️ ДАЛЕЕ", callback_data="to_branded_sections")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    total_slots = len(selected_slots)
    total_outputs_per_day = total_slots * len(selected_radios)
    total_outputs_period = total_outputs_per_day * campaign_days
    
    stations_text = "📻 ВЫБРАНЫ СТАНЦИИ:\n" + "\n".join([f"• {radio}" for radio in selected_radios])
    
    text = (
        f"● ● ● ● ○ ○ ○   [3/7] ВЫБОР ВРЕМЕНИ ВЫХОДА\n\n"
        f"⏱️ Этот шаг займет ~30 секунд\n\n"
        f"{stations_text}\n"
        f"📅 ПЕРИОД: {context.user_data.get('start_date')} - {context.user_data.get('end_date')} ({campaign_days} дней)\n\n"
        f"🕒 ВЫБЕРИТЕ ВРЕМЯ ВЫХОДА РОЛИКОВ\n\n"
        f"📊 Статистика выбора:\n"
        f"• Выбрано слотов: {total_slots}\n"
        f"• Выходов в день на всех радио: {total_outputs_per_day}\n"
        f"• Всего выходов за период: {format_number(total_outputs_period)}\n\n"
        f"✅ Готово! Следующий шаг: брендированные рубрики (15 сек)"
    )
    
    await update.message.reply_text(text, reply_markup=reply_markup)
    return TIME_SLOTS

async def time_slots(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ШАГ 3/7 - ВЫБОР ВРЕМЕНИ ВЫХОДА"""
    query = update.callback_query
    await query.answer()
    
    selected_slots = context.user_data.get("selected_time_slots", [])
    selected_radios = context.user_data.get("selected_radios", [])
    campaign_days = context.user_data.get("campaign_days", 30)
    
    keyboard = []
    keyboard.append([InlineKeyboardButton("✅ ВЫБРАТЬ ВСЕ СЛОТЫ", callback_data="select_all_slots")])
    
    for i in range(15):
        slot = TIME_SLOTS_DATA[i]
        emoji = "✅" if i in selected_slots else "⚪"
        button_text = f"{emoji} {slot['time']} • {slot['label']}"
        keyboard.append([InlineKeyboardButton(button_text, callback_data=f"time_{i}")])
    
    keyboard.append([InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_dates")])
    keyboard.append([InlineKeyboardButton("➡️ ДАЛЕЕ", callback_data="to_branded_sections")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    total_slots = len(selected_slots)
    total_outputs_per_day = total_slots * len(selected_radios)
    total_outputs_period = total_outputs_per_day * campaign_days
    
    stations_text = "📻 ВЫБРАНЫ СТАНЦИИ:\n" + "\n".join([f"• {radio}" for radio in selected_radios])
    
    text = (
        f"● ● ● ● ○ ○ ○   [3/7] ВЫБОР ВРЕМЕНИ ВЫХОДА\n\n"
        f"⏱️ Этот шаг займет ~30 секунд\n\n"
        f"{stations_text}\n"
        f"📅 ПЕРИОД: {context.user_data.get('start_date')} - {context.user_data.get('end_date')} ({campaign_days} дней)\n\n"
        f"🕒 ВЫБЕРИТЕ ВРЕМЯ ВЫХОДА РОЛИКОВ\n\n"
        f"📊 Статистика выбора:\n"
        f"• Выбрано слотов: {total_slots}\n"
        f"• Выходов в день на всех радио: {total_outputs_per_day}\n"
        f"• Всего выходов за период: {format_number(total_outputs_period)}\n\n"
        f"✅ Готово! Следующий шаг: брендированные рубрики (15 сек)"
    )
    
    await query.edit_message_text(text, reply_markup=reply_markup)
    return TIME_SLOTS

async def handle_time_slots(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ОБРАБОТЧИК ВЫБОРА ВРЕМЕНИ"""
    query = update.callback_query
    await query.answer()
    
    if query.data == "back_to_dates":
        return await campaign_dates(update, context)
    
    elif query.data == "select_all_slots":
        context.user_data["selected_time_slots"] = list(range(15))
        return await time_slots(update, context)
    
    elif query.data.startswith("time_"):
        slot_index = int(query.data.split("_")[1])
        selected_slots = context.user_data.get("selected_time_slots", [])
        
        if slot_index in selected_slots:
            selected_slots.remove(slot_index)
        else:
            selected_slots.append(slot_index)
        
        context.user_data["selected_time_slots"] = selected_slots
        return await time_slots(update, context)
    
    elif query.data == "to_branded_sections":
        if not context.user_data.get("selected_time_slots"):
            await query.answer("❌ Выберите хотя бы один временной слот!", show_alert=True)
            return TIME_SLOTS
        return await branded_sections(update, context)
    
    return TIME_SLOTS

async def branded_sections(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ШАГ 4/7 - БРЕНДИРОВАННЫЕ РУБРИКИ"""
    query = update.callback_query
    await query.answer()
    
    selected_branded = context.user_data.get("branded_section")
    
    keyboard = [
        [InlineKeyboardButton("✅ АВТОРУБРИКИ" if selected_branded == "auto" else "⚪ АВТОРУБРИКИ", callback_data="branded_auto")],
        [InlineKeyboardButton("✅ НЕДВИЖИМОСТЬ" if selected_branded == "realty" else "⚪ НЕДВИЖИМОСТЬ", callback_data="branded_realty")],
        [InlineKeyboardButton("✅ МЕДИЦИНСКИЕ" if selected_branded == "medical" else "⚪ МЕДИЦИНСКИЕ", callback_data="branded_medical")],
        [InlineKeyboardButton("✅ ИНДИВИДУАЛЬНАЯ" if selected_branded == "custom" else "⚪ ИНДИВИДУАЛЬНАЯ", callback_data="branded_custom")],
        [InlineKeyboardButton("📋 Посмотреть пример", callback_data="show_example")],
        [InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_time")],
        [InlineKeyboardButton("➡️ ДАЛЕЕ", callback_data="to_campaign_creator")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    text = (
        f"● ● ● ● ● ○ ○   [4/7] БРЕНДИРОВАННЫЕ РУБРИКИ\n\n"
        f"⏱️ Этот шаг займет ~15 секунд\n\n"
        "🎙️ ВЫБЕРИТЕ ТИП РУБРИКИ:\n\n"
        f"{'✅' if selected_branded == 'auto' else '⚪'} АВТОРУБРИКИ\n"
        "Готовые сценарии для автосалонов\n"
        "\"30 секунд о китайских автомобилях\"\n"
        "\"30 секунд об АвтоВАЗе\"\n"
        "+20% к стоимости кампании\n\n"
        f"{'✅' if selected_branded == 'realty' else '⚪'} НЕДВИЖИМОСТЬ\n"
        "Рубрики для агентств недвижимости\n"
        "\"Совет по недвижимости\"\n"
        "\"Полезно знать при покупке квартиры\"\n"
        "+15% к стоимости кампании\n\n"
        f"{'✅' if selected_branded == 'medical' else '⚪'} МЕДИЦИНСКИЕ РУБРИКИ\n"
        "Экспертные форматы для клиник\n"
        "\"Здоровое серде\"\n"
        "\"Совет врача\"\n"
        "+25% к стоимости кампании\n\n"
        f"{'✅' if selected_branded == 'custom' else '⚪'} ИНДИВИДУАЛЬНАЯ РУБРИКА\n"
        "Разработка под ваш бизнес\n"
        "Уникальный контент и сценарий\n"
        "+30% к стоимости кампании\n\n"
        f"✅ Готово! Следующий шаг: создание ролика (60 сек)"
    )
    
    await query.edit_message_text(text, reply_markup=reply_markup)
    return BRANDED_SECTIONS

async def handle_branded_sections(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ОБРАБОТЧИК БРЕНДИРОВАННЫХ РУБРИК"""
    query = update.callback_query
    await query.answer()
    
    if query.data == "back_to_time":
        return await time_slots(update, context)
    
    elif query.data == "show_example":
        example_text = (
            "Рубрики «30 секунд об АвтоВАЗе» \n"
            "Готовый комплект рубрик для АвтоВАЗа (понедельник–воскресенье). \n\n"
            "Понедельник — Интересный факт\n"
            "ВАЗ-2106, знаменитая «шестёрка», стала одной из самых массовых моделей АвтоВАЗа. За 30 лет выпуска было произведено более 4 миллионов автомобилей — рекорд для отечественного автопрома!\n\n"
            "Вторник — Интересный факт\n"
            "LADA Kalina, появившаяся в 2004 году, стала первой моделью АвтоВАЗа, оснащённой системой ABS и подушками безопасности. Именно с неё начался новый этап в развитии безопасности российских автомобилей.\n\n"
            "Среда — Интересный факт\n"
            "LADA Priora долгое время была выбором молодых водителей. За время выпуска с 2007 по 2018 год с конвейера сошло более 1 миллиона машин, а многие до сих пор на дорогах.\n\n"
            "Четверг — Интересный факт\n"
            "В 2018 году АвтоВАЗ начал экспорт LADA Vesta и LADA Largus в Европу. Эти модели хорошо зарекомендовали себя благодаря надёжности и доступной цене.\n\n"
            "Пятница — Интересный факт\n"
            "На заводе АвтоВАЗа в Тольятти работает более 30 тысяч сотрудников. Это один из крупнейших работодателей Самарской области, а сам завод называют «городом в городе».\n\n"
            "Суббота — Интересный факт\n"
            "LADA Niva не раз участвовала в ралли «Париж — Дакар». В 1980-х эта модель удивляла мир своей проходимостью и выносливостью, соревнуясь с лучшими внедорожниками мира.\n\n"
            "Воскресенье — Интересный факт\n"
            "В 2021 году LADA стала маркой №1 на российском рынке: её доля составила более 20% всех проданных автомобилей в стране. Это подтверждает доверие миллионов водителей."
        )
        
        keyboard = [[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_branded")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(example_text, reply_markup=reply_markup)
        return BRANDED_SECTIONS
    
    elif query.data == "back_to_branded":
        return await branded_sections(update, context)
    
    branded_data = {
        "branded_auto": "auto",
        "branded_realty": "realty",
        "branded_medical": "medical",
        "branded_custom": "custom"
    }
    
    if query.data in branded_data:
        context.user_data["branded_section"] = branded_data[query.data]
        return await branded_sections(update, context)
    
    elif query.data == "to_campaign_creator":
        return await campaign_creator(update, context)
    
    return BRANDED_SECTIONS

async def campaign_creator(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ШАГ 5/7 - СОЗДАНИЕ РОЛИКА (ИСПРАВЛЕННАЯ ВЕРСИЯ)"""
    query = update.callback_query
    await query.answer()
    
    base_price, discount, final_price, total_reach, daily_coverage, spots_per_day, total_coverage_percent = calculate_campaign_price_and_reach(context.user_data)
    context.user_data["base_price"] = base_price
    context.user_data["discount"] = discount
    context.user_data["final_price"] = final_price
    
    campaign_text = context.user_data.get("campaign_text", "")
    
    keyboard = [
        [InlineKeyboardButton("📝 ВВЕСТИ ТЕКСТ РОЛИКА", callback_data="enter_text")],
        [InlineKeyboardButton("🎵 ПРИШЛЮ СВОЙ РОЛИК", callback_data="provide_own_audio")],
        [InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_branded")]
    ]
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    text = (
        f"● ● ● ● ● ● ○   [5/7] СОЗДАНИЕ РОЛИКА\n\n"
        f"⏱️ Этот шаг займет ~60 секунд\n\n"
        "📝 ВАШ ТЕКСТ ДЛЯ РОЛИКА (до 500 знаков):\n\n"
        f"{campaign_text if campaign_text else '[Ваш текст появится здесь]'}\n\n"
        f"📊 Выходов в день: {spots_per_day}\n"
        f"📈 Суммарный охват слотов: {total_coverage_percent}%\n\n"
        f"💰 Предварительная стоимость:\n"
        f"   Базовая: {format_number(base_price)}₽\n"
        f"   Скидка 50%: -{format_number(discount)}₽\n"
        f"   Итоговая: {format_number(final_price)}₽\n\n"
        f"📊 Примерный охват кампании:\n"
        f"   ~{format_number(total_reach)} человек за период\n\n"
        f"✅ Готово! Следующий шаг: контактные данные (30 сек)"
    )
    
    await query.edit_message_text(text, reply_markup=reply_markup)
    return CAMPAIGN_CREATOR

async def enter_campaign_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ЭКРАН ВВОДА ТЕКСТА РОЛИКА"""
    query = update.callback_query
    await query.answer()
    
    keyboard = [
        [InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_creator")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        "📝 Введите текст для радиоролика (до 500 знаков):\n\n"
        "Пример:\n"
        "\"Автомобили в Тюмени! Новые модели в наличии. Выгодный трейд-ин и кредит 0%. "
        "Тест-драйв в день обращения!\"\n\n"
        "Отправьте текст сообщением:",
        reply_markup=reply_markup
    )
    
    return "WAITING_TEXT"

async def process_campaign_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ОБРАБОТКА ТЕКСТА РОЛИКА"""
    try:
        text = update.message.text.strip()
        if len(text) > 500:
            await update.message.reply_text(
                "❌ Текст превышает 500 знаков. Сократите текст и отправьте снова:",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_creator")]
                ])
            )
            return "WAITING_TEXT"
        
        context.user_data["campaign_text"] = text
        context.user_data["provide_own_audio"] = False
        
        char_count = len(text)
        if char_count <= 75:
            estimated_duration = 15
        elif char_count <= 100:
            estimated_duration = 20
        else:
            estimated_duration = 25
        context.user_data["duration"] = estimated_duration
        
        return await production_option(update, context)
        
    except Exception as e:
        logger.error(f"Ошибка в process_campaign_text: {e}")
        await update.message.reply_text("❌ Произошла ошибка. Попробуйте еще раз: /start")
        return ConversationHandler.END

async def enter_duration(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ЭКРАН ВВОДА ХРОНОМЕТРАЖА"""
    query = update.callback_query
    await query.answer()
    
    keyboard = [
        [InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_creator")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        "⏱️ Введите длительность ролика в секундах (10-25):\n\n"
        "Отправьте число от 10 до 25:",
        reply_markup=reply_markup
    )
    
    return "WAITING_DURATION"

async def process_duration(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ОБРАБОТКА ХРОНОМЕТРАЖА"""
    try:
        duration_text = update.message.text.strip()
        duration = int(duration_text)
        
        if duration < 10 or duration > 25:
            await update.message.reply_text(
                "❌ Длительность должна быть от 10 до 25 секунд. Попробуйте еще раз:",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_creator")]
                ])
            )
            return "WAITING_DURATION"
        
        context.user_data["duration"] = duration
        return await contact_info_from_message(update, context)
        
    except ValueError:
        await update.message.reply_text(
            "❌ Пожалуйста, введите число от 10 до 25:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_creator")]
            ])
        )
        return "WAITING_DURATION"  # ← ЭТО ИСПРАВЛЕНИЕ
        }  # конец process_duration

# ← ДОБАВИТЬ ЗДЕСЬ ↓

async def contact_info_from_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """КОНТАКТНЫЕ ДАННЫЕ ИЗ СООБЩЕНИЯ"""
    base_price, discount, final_price, total_reach, daily_coverage, spots_per_day, total_coverage_percent = calculate_campaign_price_and_reach(context.user_data)
    
    keyboard = [[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_production")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    context.user_data["current_contact_field"] = "name"
    
    text = (
        f"✨ ВАШ МЕДИАПЛАН СОСТАВЛЕН!\n\n"
        f"Реклама будет работать на {format_number(total_reach)} человек\n"
        f"Стоимость со скидкой: {format_number(final_price)}₽\n\n"
        f"──────────────────\n"
        f"👤 КАК ВАС ЗОВУТ?\n"
        f"──────────────────\n"
        f"Напишите ваше имя для оформления:"
    )
    
    await update.message.reply_text(text, reply_markup=reply_markup)
    return CONTACT_INFO
# потом идет async def production_option...

async def production_option(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ШАГ 6/7 - ПРОИЗВОДСТВО РОЛИКА"""
    query = update.callback_query if hasattr(update, "callback_query") else None
    
    if query:
        await query.answer()
    
    selected_production = context.user_data.get("production_option")
    campaign_text = context.user_data.get("campaign_text", "")
    
    keyboard = []
    
    for key, option in PRODUCTION_OPTIONS.items():
        is_selected = "✅" if selected_production == key else "⚪"
        keyboard.append([
            InlineKeyboardButton(
                f"{is_selected} {option['name']} - от {format_number(option['price'])}₽", 
                callback_data=f"production_{key}"
            )
        ])
    
    keyboard.append([InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_creator")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    text = (
        f"● ● ● ● ● ● ○   [6/7] ПРОИЗВОДСТВО РОЛИКА\n\n"
        f"⏱️ Этот шаг займет ~15 секунд\n\n"
        f"📝 ВАШ ТЕКСТ:\n\"{campaign_text[:100]}{'...' if len(campaign_text) > 100 else ''}\"\n\n"
        f"⏱️ Длительность: {context.user_data.get('duration', 20)} секунд\n\n"
        "🎙️ ВЫБЕРИТЕ ВАРИАНТ РОЛИКА:\n\n"
        "⚪ СТАНДАРТНЫЙ РОЛИК - от 2 000₽\n"
        "• Профессиональная озвучка\n• Музыкальное оформление\n• Срок: 2-3 дня\n\n"
        "⚪ ПРЕМИУМ РОЛИК - от 5 000₽\n"
        "• Озвучка 2-мя голосами\n• Индивидуальная музыка\n• Срочное производство 1 день\n"
    )
    
    text += f"\n✅ Готово! Следующий шаг: контактные данные (30 сек)"
    
    if query:
        await query.edit_message_text(text, reply_markup=reply_markup)
    else:
        await update.message.reply_text(text, reply_markup=reply_markup)
    
    return PRODUCTION_OPTION

async def handle_production_option(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ОБРАБОТЧИК ВЫБОРА ПРОИЗВОДСТВА"""
    query = update.callback_query
    await query.answer()
    
    if query.data == "back_to_creator":
        return await campaign_creator(update, context)
    
    elif query.data.startswith("production_"):
        production_key = query.data.replace("production_", "")
        if production_key in PRODUCTION_OPTIONS:
            context.user_data["production_option"] = production_key
            context.user_data["production_cost"] = PRODUCTION_OPTIONS[production_key]["price"]
            # АВТОМАТИЧЕСКИЙ ПЕРЕХОД К КОНТАКТАМ
            return await contact_info(update, context)
    
    return PRODUCTION_OPTION

async def contact_info(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ШАГ 7/7 - КОНТАКТНЫЕ ДАННЫЕ"""
    query = update.callback_query
    await query.answer()
    
    base_price, discount, final_price, total_reach, daily_coverage, spots_per_day, total_coverage_percent = calculate_campaign_price_and_reach(context.user_data)
    
    keyboard = [[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_production")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    context.user_data["current_contact_field"] = "name"
    
    text = (
        f"✨ ВАШ МЕДИАПЛАН СОСТАВЛЕН!\n\n"
        f"Реклама будет работать на {format_number(total_reach)} человек\n"
        f"Стоимость со скидкой: {format_number(final_price)}₽\n\n"
        f"──────────────────\n"
        f"👤 КАК ВАС ЗОВУТ?\n"
        f"──────────────────\n"
        f"Напишите ваше имя для оформления:"
    )
    
    await query.edit_message_text(text, reply_markup=reply_markup)
    return CONTACT_INFO

async def process_contact_info(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ОБНОВЛЕННАЯ ОБРАБОТКА КОНТАКТНЫХ ДАННЫХ С НАВИГАЦИЕЙ НАЗАД"""
    try:
        text = update.message.text.strip()
        current_field = context.user_data.get("current_contact_field", "name")
        
        if current_field == "name":
            context.user_data["contact_name"] = text
            context.user_data["current_contact_field"] = "phone"
            
            keyboard = [[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_contact_name")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await update.message.reply_text(
                "📞 Введите ваш телефон:\n\n"
                "Пример: +79123456789 или любой формат",
                reply_markup=reply_markup
            )
            return CONTACT_INFO
            
        elif current_field == "phone":
            context.user_data["phone"] = text
            context.user_data["current_contact_field"] = "email"
            
            keyboard = [[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_contact_phone")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await update.message.reply_text(
                "📧 Введите ваш email:",
                reply_markup=reply_markup
            )
            return CONTACT_INFO
            
        elif current_field == "email":
            context.user_data["email"] = text
            context.user_data["current_contact_field"] = "company"
            
            keyboard = [[InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_contact_email")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await update.message.reply_text(
                "🏢 Введите название компании:",
                reply_markup=reply_markup
            )
            return CONTACT_INFO
            
        elif current_field == "company":
            context.user_data["company"] = text
            context.user_data.pop("current_contact_field", None)
            return await show_confirmation_from_message(update, context)
            
    except Exception as e:
        logger.error(f"Ошибка в process_contact_info: {e}")
        await update.message.reply_text(
            "❌ Произошла ошибка. Пожалуйста, начните заново: /start"
        )
        return ConversationHandler.END

async def show_confirmation_from_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ПОДТВЕРЖДЕНИЕ ЗАЯВКИ ИЗ СООБЩЕНИЯ"""
    base_price, discount, final_price, total_reach, daily_coverage, spots_per_day, total_coverage_percent = calculate_campaign_price_and_reach(context.user_data)
    
    stations_text = ""
    for radio in context.user_data.get("selected_radios", []):
        listeners = STATION_COVERAGE.get(radio, 0)
        stations_text += f"• {radio}: ~{format_number(listeners)} слушателей в день\n"
    
    slots_text, total_slots_coverage = get_time_slots_detailed_text(context.user_data.get("selected_time_slots", []))
    
    confirmation_text = f"""
🎉 ВСЕ ГОТОВО! ПОДТВЕРЖДЕНИЕ ЗАЯВКИ

👤 ВАШИ ДАННЫЕ:
Имя: {context.user_data.get('contact_name', 'Не указано')}
Телефон: {context.user_data.get('phone', 'Не указан')}
Email: {context.user_data.get('email', 'Не указан')}
Компания: {context.user_data.get('company', 'Не указана')}

📊 ПАРАМЕТРЫ КАМПАНИИ:

📻 РАДИОСТАНЦИИ:
{stations_text}
📅 ПЕРИОД: {context.user_data.get('start_date')} - {context.user_data.get('end_date')} ({context.user_data.get('campaign_days')} дней)

🕒 ВЫБРАННЫЕ ВРЕМЕННЫЕ СЛОТЫ:
{slots_text}
• Суммарный охват слотов: {total_slots_coverage}%

🎙️ РУБРИКА: {get_branded_section_name(context.user_data.get('branded_section'))}
⏱️ РОЛИК: {PRODUCTION_OPTIONS.get(context.user_data.get('production_option', 'ready'), {}).get('name', 'Не выбрано')}
📏 ХРОНОМЕТРАЖ: {context.user_data.get('duration', 20)} сек

🎯 РАСЧЕТНЫЙ ОХВАТ:
• Выходов в день: {spots_per_day}
• Всего выходов: {spots_per_day * context.user_data.get('campaign_days', 30)}
• Уникальных слушателей в день: ~{format_number(daily_coverage)} чел.
• Общий охват за период: ~{format_number(total_reach)} чел.

💰 СТОИМОСТЬ:
Базовая: {format_number(base_price)}₽
Скидка 50%: -{format_number(discount)}₽
Итоговая: {format_number(final_price)}₽
"""
    
    keyboard = [
        [InlineKeyboardButton("📤 ОТПРАВИТЬ ЗАЯВКУ", callback_data="submit_campaign")],
        [InlineKeyboardButton("◀️ ВЕРНУТЬСЯ К ВЫБОРУ РАДИО", callback_data="back_to_radio")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(confirmation_text, reply_markup=reply_markup)
    return CONFIRMATION

async def show_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ПОДТВЕРЖДЕНИЕ ЗАЯВКИ"""
    query = update.callback_query
    await query.answer()
    
    base_price, discount, final_price, total_reach, daily_coverage, spots_per_day, total_coverage_percent = calculate_campaign_price_and_reach(context.user_data)
    
    stations_text = ""
    for radio in context.user_data.get("selected_radios", []):
        listeners = STATION_COVERAGE.get(radio, 0)
        stations_text += f"• {radio}: ~{format_number(listeners)} слушателей в день\n"
    
    slots_text, total_slots_coverage = get_time_slots_detailed_text(context.user_data.get("selected_time_slots", []))
    
    confirmation_text = f"""
🎉 ВСЕ ГОТОВО! ПОДТВЕРЖДЕНИЕ ЗАЯВКИ

👤 ВАШИ ДАННЫЕ:
Имя: {context.user_data.get('contact_name', 'Не указано')}
Телефон: {context.user_data.get('phone', 'Не указан')}
Email: {context.user_data.get('email', 'Не указан')}
Компания: {context.user_data.get('company', 'Не указана')}

📊 ПАРАМЕТРЫ КАМПАНИИ:

📻 РАДИОСТАНЦИИ:
{stations_text}
📅 ПЕРИОД: {context.user_data.get('start_date')} - {context.user_data.get('end_date')} ({context.user_data.get('campaign_days')} дней)

🕒 ВЫБРАННЫЕ ВРЕМЕННЫЕ СЛОТЫ:
{slots_text}
• Суммарный охват слотов: {total_slots_coverage}%

🎙️ РУБРИКА: {get_branded_section_name(context.user_data.get('branded_section'))}
⏱️ РОЛИК: {PRODUCTION_OPTIONS.get(context.user_data.get('production_option', 'ready'), {}).get('name', 'Не выбрано')}
📏 ХРОНОМЕТРАЖ: {context.user_data.get('duration', 20)} сек

🎯 РАСЧЕТНЫЙ ОХВАТ:
• Выходов в день: {spots_per_day}
• Всего выходов: {spots_per_day * context.user_data.get('campaign_days', 30)}
• Уникальных слушателей в день: ~{format_number(daily_coverage)} чел.
• Общий охват за период: ~{format_number(total_reach)} чел.

💰 СТОИМОСТЬ:
Базовая: {format_number(base_price)}₽
Скидка 50%: -{format_number(discount)}₽
Итоговая: {format_number(final_price)}₽
"""
    
    keyboard = [
        [InlineKeyboardButton("📤 ОТПРАВИТЬ ЗАЯВКУ", callback_data="submit_campaign")],
        [InlineKeyboardButton("◀️ ВЕРНУТЬСЯ К ВЫБОРУ РАДИО", callback_data="back_to_radio")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(confirmation_text, reply_markup=reply_markup)
    return CONFIRMATION

async def handle_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ОБРАБОТКА ПОДТВЕРЖДЕНИЯ"""
    query = update.callback_query
    await query.answer()
    
    if query.data == "back_to_radio":
        saved_contacts = {
            "contact_name": context.user_data.get("contact_name"),
            "phone": context.user_data.get("phone"), 
            "email": context.user_data.get("email"),
            "company": context.user_data.get("company")
        }
        context.user_data.clear()
        context.user_data.update(saved_contacts)
        return await radio_selection(update, context)
    
    elif query.data == "submit_campaign":
        try:
            if not check_rate_limit(query.from_user.id):
                await query.answer(
                    "❌ Вы превысили лимит в 5 заявок в день. Попробуйте завтра или свяжитесь с поддержкой: @AlexeyKhlistunov",
                    show_alert=True
                )
                return CONFIRMATION
            
            base_price, discount, final_price, total_reach, daily_coverage, spots_per_day, total_coverage_percent = calculate_campaign_price_and_reach(context.user_data)
            
            campaign_number = f"R-{datetime.now().strftime('%H%M%S')}"
            conn = sqlite3.connect("campaigns.db")
            cursor = conn.cursor()
            
            cursor.execute("""
                INSERT INTO campaigns 
                (user_id, campaign_number, radio_stations, start_date, end_date, campaign_days,
                 time_slots, branded_section, campaign_text, production_option, contact_name,
                 company, phone, email, duration, base_price, discount, final_price, actual_reach, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                query.from_user.id,
                campaign_number,
                ",".join(context.user_data.get("selected_radios", [])),
                context.user_data.get("start_date"),
                context.user_data.get("end_date"),
                context.user_data.get("campaign_days"),
                ",".join(map(str, context.user_data.get("selected_time_slots", []))),
                context.user_data.get("branded_section", ""),
                context.user_data.get("campaign_text", ""),
                context.user_data.get("production_option", ""),
                context.user_data.get("contact_name", ""),
                context.user_data.get("company", ""),
                context.user_data.get("phone", ""),
                context.user_data.get("email", ""),
                context.user_data.get("duration", 20),
                base_price,
                discount,
                final_price,
                total_reach,  # СОХРАНЯЕМ ОХВАТ В БД
                "active"
            ))
            
            conn.commit()
            conn.close()
            
            await send_admin_notification(context, context.user_data, campaign_number)
            
            success_text = f"""
🎉 ЗАЯВКА ПРИНЯТА! 

Спасибо, что выбрали нас! Ваша реклама скоро выйдет в эфир 🎙️

📋 № заявки: {campaign_number}
📅 Старт кампании: в течение 3 дней
💰 Ваша цена: {format_number(final_price)}₽ (скидка 50%)
📊 Охват аудитории: ~{format_number(total_reach)} человек за период

Выберите дальнейшее действие:
"""
            
            keyboard = [
                [InlineKeyboardButton("📊 EXCEL МЕДИАПЛАН", callback_data="generate_excel")],
                [InlineKeyboardButton("📞 СВЯЗЬ С МЕНЕДЖЕРОМ", callback_data="contact_manager")],
                [InlineKeyboardButton("🚀 НОВАЯ КАМПАНИЯ", callback_data="new_campaign")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await query.message.reply_text(success_text, reply_markup=reply_markup)
            return FINAL_ACTIONS
            
        except Exception as e:
            logger.error(f"Ошибка при сохранении заявки: {e}")
            await query.message.reply_text(
                "❌ Произошла ошибка при сохранении заявки.\n"
                "Пожалуйста, начните заново: /start\n"
                "Или свяжитесь с поддержкой: t.me/AlexeyKhlistunov"
            )
            return ConversationHandler.END
    
    return CONFIRMATION

async def handle_final_actions(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ОБРАБОТКА ФИНАЛЬНЫХ ДЕЙСТВИЙ"""
    try:
        query = update.callback_query
        await query.answer()
        
        if query.data == "generate_excel":
            user_id = query.from_user.id
            conn = sqlite3.connect("campaigns.db")
            cursor = conn.cursor()
            cursor.execute("SELECT campaign_number FROM campaigns WHERE user_id = ? ORDER BY created_at DESC LIMIT 1", (user_id,))
            campaign_data = cursor.fetchone()
            conn.close()
            
            if campaign_data:
                campaign_number = campaign_data[0]
                try:
                    excel_buffer = create_excel_file_from_db(campaign_number)
                    if excel_buffer:
                        await query.message.reply_document(
                            document=excel_buffer,
                            filename=f"mediaplan_{campaign_number}.xlsx",
                            caption=f"📊 Ваш медиаплан кампании #{campaign_number}"
                        )
                    else:
                        await query.message.reply_text("❌ Ошибка при создании Excel. Попробуйте еще раз.")
                except Exception as e:
                    logger.error(f"Ошибка Excel: {e}")
                    await query.message.reply_text("❌ Ошибка при создании Excel. Попробуйте еще раз.")
            else:
                await query.message.reply_text("❌ Не найдено данных о кампании.")
            return FINAL_ACTIONS
        
        elif query.data == "contact_manager":
            await query.message.reply_text(
                "📞 Свяжитесь с менеджером:\n"
                "👤 @AlexeyKhlistunov\n"
                "📱 +7 (922) 044-66-44\n"
                "✉️ a.khlistunov@gmail.com"
            )
            return FINAL_ACTIONS
        
        elif query.data == "new_campaign":
            saved_contacts = {
                "contact_name": context.user_data.get("contact_name"),
                "phone": context.user_data.get("phone"),
                "email": context.user_data.get("email"),
                "company": context.user_data.get("company")
            }
            context.user_data.clear()
            context.user_data.update(saved_contacts)
            await query.message.reply_text("🚀 Начинаем новую кампанию!")
            return await radio_selection(update, context)
        
        return FINAL_ACTIONS
        
    except Exception as e:
        logger.error(f"Ошибка в handle_final_actions: {e}")
        await query.message.reply_text("❌ Ошибка. Начните заново: /start")
        return ConversationHandler.END

async def personal_cabinet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ЛИЧНЫЙ КАБИНЕТ"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    conn = sqlite3.connect("campaigns.db")
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT campaign_number, start_date, end_date, final_price, actual_reach 
        FROM campaigns 
        WHERE user_id = ? AND status = 'active'
        ORDER BY created_at DESC LIMIT 5
    """, (user_id,))
    active_orders = cursor.fetchall()
    
    cursor.execute("""
        SELECT campaign_number, start_date, end_date, final_price, actual_reach 
        FROM campaigns 
        WHERE user_id = ? AND status = 'completed'
        ORDER BY created_at DESC LIMIT 5
    """, (user_id,))
    completed_orders = cursor.fetchall()
    
    cursor.execute("""
        SELECT COUNT(*), SUM(final_price), SUM(actual_reach)
        FROM campaigns 
        WHERE user_id = ? AND strftime('%Y', created_at) = '2025'
    """, (user_id,))
    stats = cursor.fetchone()
    conn.close()
    
    orders_text = "📋 ЛИЧНЫЙ КАБИНЕТ\n\n"
    
    if active_orders:
        orders_text += "🚀 АКТИВНЫЕ КАМПАНИИ\n"
        for order in active_orders:
            orders_text += f"• {order[0]} | {order[1]}-{order[2]} | {format_number(order[3])}₽ | {format_number(order[4] or 0)} охват\n"
        orders_text += "\n"
    
    if completed_orders:
        orders_text += "📊 ЗАВЕРШЕННЫЕ КАМПАНИИ\n"
        for order in completed_orders:
            orders_text += f"• {order[0]} | {order[1]}-{order[2]} | {format_number(order[3])}₽ | {format_number(order[4] or 0)} охват\n"
        orders_text += "\n"
    
    if stats and stats[0]:
        orders_text += f"💎 СТАТИСТИКА ЗА 2025:\n"
        orders_text += f"• {stats[0]} кампаний | {format_number(stats[1] or 0)}₽ | {format_number(stats[2] or 0)} охват\n"
    
    keyboard = [
        [InlineKeyboardButton("📈 ДЕТАЛЬНАЯ СТАТИСТИКА", callback_data="detailed_stats")],
        [InlineKeyboardButton("🚀 НОВАЯ КАМПАНИЯ", callback_data="new_campaign")],
        [InlineKeyboardButton("🔙 НАЗАД", callback_data="back_to_main")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(orders_text, reply_markup=reply_markup)
    return MAIN_MENU

async def detailed_statistics(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ДЕТАЛЬНАЯ СТАТИСТИКА"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    conn = sqlite3.connect("campaigns.db")
    cursor = conn.cursor()
    
    # Все кампании пользователя
    cursor.execute("""
        SELECT campaign_number, start_date, end_date, final_price, actual_reach, status, created_at
        FROM campaigns 
        WHERE user_id = ?
        ORDER BY created_at DESC
    """, (user_id,))
    all_campaigns = cursor.fetchall()
    
    # Статистика по годам
    cursor.execute("""
        SELECT strftime('%Y', created_at) as year, 
               COUNT(*), SUM(final_price), SUM(actual_reach)
        FROM campaigns 
        WHERE user_id = ?
        GROUP BY strftime('%Y', created_at)
        ORDER BY year DESC
    """, (user_id,))
    yearly_stats = cursor.fetchall()
    
    conn.close()
    
    stats_text = "📈 ДЕТАЛЬНАЯ СТАТИСТИКА\n\n"
    
    if all_campaigns:
        stats_text += "📋 ВСЕ КАМПАНИИ:\n"
        for campaign in all_campaigns:
            status_emoji = "🟢" if campaign[5] == "active" else "🔴"
            stats_text += f"{status_emoji} {campaign[0]} | {campaign[1]}-{campaign[2]} | {format_number(campaign[3])}₽ | {format_number(campaign[4] or 0)} охват\n"
        stats_text += "\n"
    
    if yearly_stats:
        stats_text += "📊 СТАТИСТИКА ПО ГОДАМ:\n"
        for year_stat in yearly_stats:
            stats_text += f"• {year_stat[0]} год: {year_stat[1]} кампаний | {format_number(year_stat[2] or 0)}₽ | {format_number(year_stat[3] or 0)} охват\n"
        stats_text += "\n"
    
    if not all_campaigns:
        stats_text += "📭 У вас пока нет завершенных кампаний\n"
    
    keyboard = [
        [InlineKeyboardButton("🔙 В ЛИЧНЫЙ КАБИНЕТ", callback_data="personal_cabinet")],
        [InlineKeyboardButton("🚀 НОВАЯ КАМПАНИЯ", callback_data="new_campaign")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(stats_text, reply_markup=reply_markup)
    return MAIN_MENU

async def statistics(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ВОЗРАСТНАЯ СТРУКТУРА"""
    query = update.callback_query
    await query.answer()
    
    keyboard = [
        [InlineKeyboardButton("🎯 ПОДОБРАТЬ ОПТИМАЛЬНЫЙ МИКС", callback_data="create_campaign")],
        [InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_main")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        "📊 ВОЗРАСТНАЯ СТРУКТУРА \n"
        "ЯЛУТОРОВСКА И ЗАВОДОУКОВСКА\n"
        "— КЛЮЧЕВОЙ ФАКТОР ВАШЕГО ОХВАТА\n\n"
        "🎯 ПОНИМАЕМ РЕАЛЬНУЮ АУДИТОРИЮ:\n\n"
        "Демография городов определяет \n"
        "эффективность каждой радиостанции\n\n"
        "📈 НАШ АНАЛИЗ ПОКАЗЫВАЕТ:\n\n"
        "• Ялуторовск — выше доля 65+\n"
        "• Заводоуковск — моложе Ялуторовска,\n"
        "  но демография смещена в сторону\n"
        "  старших возрастов\n"
        "• Общий тренд — отток молодежи + высокая\n"
        "  доля старшего поколения\n\n"
        "📻 РЕАЛЬНЫЕ ЦИФРЫ ОХВАТА\n"
        "С УЧЕТОМ ВОЗРАСТНОЙ СТРУКТУРЫ:\n\n"
        "🎵 РЕТРО FM — 3,200-4,000 слушателей\n"
        "   Идеально для аудитории 35-65 лет\n\n"
        "🚗 АВТОРАДИО — 2,900-3,600 слушателей  \n"
        "   Стабильный охват автомобилистов 25-55\n\n"
        "💖 LOVE RADIO — 480-600 слушателей\n"
        "   Молодежь 16-35, охват снижен на 40-50%\n\n"
        "🎭 ЮМОР FM — 1,120-1,400 слушателей\n"
        "   Смешанная аудитория 25-45 лет\n\n"
        "🏠 РАДИО ДАЧА — 2,900-3,600 слушателей\n"
        "   Семейная аудитория 35-60 лет\n\n"
        "🎸 РАДИО ШАНСОН — 2,600-3,200 слушателей\n"
        "   Мужчины 30-60+ лет\n\n"
        "💡 ВАШЕ КОНКУРЕНТНОЕ ПРЕИМУЩЕСТВО:\n"
        "Используйте станции, которые \n"
        "естественно преобладают в регионе\n"
        "— это даст максимальный охват\n"
        "при оптимальном бюджете",
        reply_markup=reply_markup
    )
    return MAIN_MENU

async def contacts_details(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """КОНТАКТЫ"""
    query = update.callback_query
    await query.answer()
    
    keyboard = [
        [InlineKeyboardButton("◀️ НАЗАД", callback_data="back_to_main")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    contacts_text = (
        "📞 КОНТАКТЫ И РЕКВИЗИТЫ\n\n"
        "📍 Ялуторовск • Заводоуковск\n"
        "📱 +7 (922) 044-66-44\n"
        "✉️ a.khlistunov@gmail.com\n"
        "👤 Telegram: @AlexeyKhlistunov\n\n"
        "Юридическая информация:\n"
        "ИП Хлыстунов Алексей Александрович\n"
        "ОГРНИП 315723200067362"
    )
    
    await query.edit_message_text(contacts_text, reply_markup=reply_markup)
    return MAIN_MENU

async def handle_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ОБНОВЛЕННЫЙ ОБРАБОТЧИК ГЛАВНОГО МЕНЮ"""
    query = update.callback_query
    await query.answer()
    
    if query.data == "create_campaign":
        saved_contacts = {
            "contact_name": context.user_data.get("contact_name"),
            "phone": context.user_data.get("phone"),
            "email": context.user_data.get("email"),
            "company": context.user_data.get("company")
        }
        context.user_data.clear()
        context.user_data.update(saved_contacts)
        return await radio_selection(update, context)
    
    elif query.data == "statistics":
        return await statistics(update, context)
    
    elif query.data == "about":
        return await about_section(update, context)
    
    elif query.data == "personal_cabinet":
        return await personal_cabinet(update, context)
    
    elif query.data == "contacts_details":
        return await contacts_details(update, context)
    
    elif query.data == "detailed_stats":
        return await detailed_statistics(update, context)
    
    elif query.data == "back_to_main":
        return await start(update, context)
    
    elif query.data == "back_to_radio":
        saved_contacts = {
            "contact_name": context.user_data.get("contact_name"),
            "phone": context.user_data.get("phone"),
            "email": context.user_data.get("email"),
            "company": context.user_data.get("company")
        }
        context.user_data.clear()
        context.user_data.update(saved_contacts)
        return await radio_selection(update, context)
    
    elif query.data == "back_to_dates":
        return await campaign_dates(update, context)
    
    elif query.data == "back_to_time":
        return await time_slots(update, context)
    
    elif query.data == "back_to_branded":
        return await branded_sections(update, context)
    
    elif query.data == "back_to_creator":
        return await campaign_creator(update, context)
    
    elif query.data == "back_to_production":
        return await production_option(update, context)
    
    # НОВЫЕ ОБРАБОТЧИКИ ДЛЯ НАВИГАЦИИ В КОНТАКТАХ
    elif query.data == "back_to_contact_name":
        context.user_data["current_contact_field"] = "name"
        return await contact_info(update, context)
    
    elif query.data == "back_to_contact_phone":
        context.user_data["current_contact_field"] = "phone"
        return await contact_info(update, context)
    
    elif query.data == "back_to_contact_email":
        context.user_data["current_contact_field"] = "email"
        return await contact_info(update, context)
    
    # ИСПРАВЛЕННЫЙ ОБРАБОТЧИК ДЛЯ "ПРИШЛЮ СВОЙ РОЛИК"
    elif query.data == "provide_own_audio":
        print("🔔 КНОПКА provide_own_audio НАЖАТА!")
        logger.info("🔔 КНОПКА provide_own_audio НАЖАТА!")
        # СРАЗУ переходим к вводу хронометража
        context.user_data["provide_own_audio"] = True
        context.user_data["campaign_text"] = ""  # Очищаем текст если был
        return await enter_duration(update, context)
    
    elif query.data == "skip_text":
        context.user_data["campaign_text"] = ""
        return await production_option(update, context)
    
    elif query.data == "cancel_text":
        return await campaign_creator(update, context)
    
    elif query.data == "cancel_duration":
        return await campaign_creator(update, context)
    
    elif query.data == "cancel_period":
        return await campaign_dates(update, context)
    
    elif query.data == "to_production_option":
        return await production_option(update, context)
    
    elif query.data == "enter_duration":
        return await enter_duration(update, context)
    
    elif query.data == "enter_text":
        return await enter_campaign_text(update, context)
    
    elif query.data == "submit_campaign":
        return await handle_confirmation(update, context)
    
    return MAIN_MENU

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ОТМЕНА"""
    await update.message.reply_text(
        "❌ Операция отменена.\n\n"
        "Для начала новой кампании используйте /start"
    )
    return ConversationHandler.END

def main():
    """ОСНОВНАЯ ФУНКЦИЯ"""
    if init_db():
        logger.info("Бот запущен успешно")
    else:
        logger.error("Ошибка инициализации БД")
    
    application = Application.builder().token(TOKEN).build()
    
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            MAIN_MENU: [
                CallbackQueryHandler(handle_main_menu, pattern="^.*$")
            ],
            RADIO_SELECTION: [
                CallbackQueryHandler(handle_radio_selection, pattern="^.*$")
            ],
            CAMPAIGN_DATES: [
                CallbackQueryHandler(handle_campaign_dates, pattern="^.*$")
            ],
            "WAITING_START_DATE": [
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_start_date),
                CallbackQueryHandler(handle_main_menu, pattern="^back_to_radio$")
            ],
            "WAITING_END_DATE": [
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_end_date),
                CallbackQueryHandler(handle_main_menu, pattern="^back_to_radio$")
            ],
            TIME_SLOTS: [
                CallbackQueryHandler(handle_time_slots, pattern="^.*$")
            ],
            BRANDED_SECTIONS: [
                CallbackQueryHandler(handle_branded_sections, pattern="^.*$")
            ],
         CAMPAIGN_CREATOR: [
    CallbackQueryHandler(handle_main_menu, pattern="^(back_to_|skip_text|cancel_text|to_production_option|enter_text|enter_duration|provide_own_audio)"),
    CallbackQueryHandler(enter_campaign_text, pattern="^enter_text$"),
    CallbackQueryHandler(enter_duration, pattern="^enter_duration$")
],
            "WAITING_TEXT": [
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_campaign_text),
                CallbackQueryHandler(handle_main_menu, pattern="^back_to_creator$"),
                CallbackQueryHandler(handle_main_menu, pattern="^cancel_text$")
            ],
            "WAITING_DURATION": [
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_duration),
                CallbackQueryHandler(handle_main_menu, pattern="^back_to_creator$"),
                CallbackQueryHandler(handle_main_menu, pattern="^cancel_duration$")
            ],
            PRODUCTION_OPTION: [
                CallbackQueryHandler(handle_production_option, pattern="^.*$")
            ],
            CONTACT_INFO: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_contact_info),
                CallbackQueryHandler(handle_main_menu, pattern="^(back_to_production|back_to_contact_name|back_to_contact_phone|back_to_contact_email)$"),
                CommandHandler("cancel", cancel)
            ],
            CONFIRMATION: [
                CallbackQueryHandler(handle_confirmation, pattern="^.*$")
            ],
            FINAL_ACTIONS: [
                CallbackQueryHandler(handle_final_actions, pattern="^.*$")
            ]
        },
        fallbacks=[CommandHandler("start", start), CommandHandler("cancel", cancel)],
        allow_reentry=True
    )
    
    application.add_handler(conv_handler)
    
    application.add_handler(CallbackQueryHandler(
        lambda update, context: update.callback_query.answer(), 
        pattern="^(call_|email_)"
    ))
    
    if "RENDER" in os.environ:
        application.run_webhook(
            listen="0.0.0.0",
            port=int(os.environ.get("PORT", 8443)),
            url_path=TOKEN,
            webhook_url=f"https://{os.environ.get('RENDER_SERVICE_NAME', 'telegram-radio-bot')}.onrender.com/{TOKEN}"
        )
    else:
        application.run_polling()

if __name__ == "__main__":
    main()

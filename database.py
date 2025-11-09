import os
import logging
import sqlite3
from datetime import datetime, timedelta
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

# Токен бота
TOKEN = "8281804030:AAEFEYgqigL3bdH4DL0zl1tW71fwwo_8cyU"

# Ваш Telegram ID для уведомлений
ADMIN_TELEGRAM_ID = 174046571

# Цены и параметры
BASE_PRICE_PER_SECOND = 2.0  # 2₽ за секунду
MIN_PRODUCTION_COST = 2000
MIN_BUDGET = 7000

# Обновленные данные слотов с разным охватом
TIME_SLOTS_DATA = [
    {"time": "06:00-07:00", "label": "Подъем, сборы", "premium": True, "coverage_percent": 6},
    {"time": "07:00-08:00", "label": "Утренние поездки", "premium": True, "coverage_percent": 10},
    {"time": "08:00-09:00", "label": "Пик трафика", "premium": True, "coverage_percent": 12},
    {"time": "09:00-10:00", "label": "Начало работы", "premium": True, "coverage_percent": 8},
    {"time": "10:00-11:00", "label": "Рабочий процесс", "premium": True, "coverage_percent": 7},
    {"time": "11:00-12:00", "label": "Предобеденное время", "premium": True, "coverage_percent": 6},
    {"time": "12:00-13:00", "label": "Обеденный перерыв", "premium": True, "coverage_percent": 5},
    {"time": "13:00-14:00", "label": "После обеда", "premium": True, "coverage_percent": 5},
    {"time": "14:00-15:00", "label": "Вторая половина дня", "premium": True, "coverage_percent": 5},
    {"time": "15:00-16:00", "label": "Рабочий финиш", "premium": True, "coverage_percent": 6},
    {"time": "16:00-17:00", "label": "Конец рабочего дня", "premium": True, "coverage_percent": 7},
    {"time": "17:00-18:00", "label": "Вечерние поездки", "premium": True, "coverage_percent": 10},
    {"time": "18:00-19:00", "label": "Пик трафика", "premium": True, "coverage_percent": 8},
    {"time": "19:00-20:00", "label": "Домашний вечер", "premium": True, "coverage_percent": 4},
    {"time": "20:00-21:00", "label": "Вечерний отдых", "premium": True, "coverage_percent": 4}
]

# Обновленные данные охвата (усредненные)
STATION_COVERAGE = {
    "LOVE RADIO": 540,
    "АВТОРАДИО": 3250,
    "РАДИО ДАЧА": 3250,
    "РАДИО ШАНСОН": 2900,
    "РЕТРО FM": 3600,
    "ЮМОР FM": 1260
}

BRANDED_SECTION_PRICES = {
    "auto": 1.2,
    "realty": 1.15,
    "medical": 1.25,
    "custom": 1.3
}

PRODUCTION_OPTIONS = {
    "standard": {"price": 2000, "name": "СТАНДАРТНЫЙ РОЛИК", "desc": "Профессиональная озвучка, музыкальное оформление, срок: 2-3 дня"},
    "premium": {"price": 5000, "name": "ПРЕМИУМ РОЛИК", "desc": "Озвучка 2-мя голосами, индивидуальная музыка, срочное производство 1 день"}
}

# Инициализация базы данных
def init_db():
    try:
        conn = sqlite3.connect("campaigns.db")
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS campaigns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                campaign_number TEXT,
                radio_stations TEXT,
                start_date TEXT,
                end_date TEXT,
                campaign_days INTEGER,
                time_slots TEXT,
                branded_section TEXT,
                campaign_text TEXT,
                production_option TEXT,
                contact_name TEXT,
                company TEXT,
                phone TEXT,
                email TEXT,
                duration INTEGER,
                base_price INTEGER,
                discount INTEGER,
                final_price INTEGER,
                actual_reach INTEGER,
                status TEXT DEFAULT "active",
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Таблица для ограничения частоты запросов
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS rate_limits (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                action_type TEXT,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        conn.commit()
        conn.close()
        logger.info("База данных инициализирована успешно")
        return True
    except Exception as e:
        logger.error(f"Ошибка инициализации БД: {e}")
        return False

def validate_phone(phone: str) -> bool:
    """Упрощенная валидация телефона"""
    if not phone:
        return False
    return True

def validate_date(date_text: str) -> bool:
    """Проверка валидности даты"""
    try:
        date = datetime.strptime(date_text, "%d.%m.%Y")
        if date < datetime.now().replace(hour=0, minute=0, second=0, microsecond=0):
            return False
        if date > datetime.now() + timedelta(days=365):
            return False
        return True
    except ValueError:
        return False

def format_number(num):
    return f"{num:,}".replace(",", " ")

def check_rate_limit(user_id: int) -> bool:
    """Проверка ограничения в 5 заявок в день"""
    try:
        conn = sqlite3.connect("campaigns.db")
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT COUNT(*) FROM campaigns 
            WHERE user_id = ? AND created_at >= datetime("now", "-1 day")
        """, (user_id,))
        
        count = cursor.fetchone()[0]
        conn.close()
        
        return count < 5
    except Exception as e:
        logger.error(f"Ошибка проверки лимитa: {e}")
        return True

def calculate_campaign_price_and_reach(user_data):
    """ОБНОВЛЕННАЯ ФУНКЦИЯ РАСЧЕТА С РАЗНЫМ ОХВАТОМ СЛОТОВ"""
    try:
        base_duration = user_data.get("duration", 20)
        campaign_days = user_data.get("campaign_days", 30)
        selected_radios = user_data.get("selected_radios", [])
        selected_time_slots = user_data.get("selected_time_slots", [])
        
        if not selected_radios or not selected_time_slots:
            return 0, 0, MIN_BUDGET, 0, 0, 0, 0
            
        num_stations = len(selected_radios)
        spots_per_day = len(selected_time_slots) * num_stations
        
        cost_per_spot = base_duration * BASE_PRICE_PER_SECOND
        base_air_cost = cost_per_spot * spots_per_day * campaign_days
        
        time_multiplier = 1.0
        for slot_index in selected_time_slots:
            if 0 <= slot_index < len(TIME_SLOTS_DATA):
                slot = TIME_SLOTS_DATA[slot_index]
                if slot["premium"]:
                    time_multiplier = max(time_multiplier, 1.1)
        
        branded_multiplier = 1.0
        branded_section = user_data.get("branded_section")
        if branded_section in BRANDED_SECTION_PRICES:
            branded_multiplier = BRANDED_SECTION_PRICES[branded_section]
        
        production_cost = user_data.get("production_cost", 0)
        air_cost = int(base_air_cost * time_multiplier * branded_multiplier)
        base_price = air_cost + production_cost
        
        discount = int(base_price * 0.5)
        discounted_price = base_price - discount
        final_price = max(discounted_price, MIN_BUDGET)
        
        # ОБНОВЛЕННЫЙ РАСЧЕТ ОХВАТА С РАЗНЫМИ % СЛОТОВ
        total_listeners = sum(STATION_COVERAGE.get(radio, 0) for radio in selected_radios)
        
        # Сумма % охвата выбранных слотов
        total_coverage_percent = 0
        for slot_index in selected_time_slots:
            if 0 <= slot_index < len(TIME_SLOTS_DATA):
                slot = TIME_SLOTS_DATA[slot_index]
                total_coverage_percent += slot["coverage_percent"]
        
        # Уникальный охват с учетом пересечения аудитории (0.7)
        unique_daily_coverage = int(total_listeners * 0.7 * (total_coverage_percent / 100))
        total_reach = int(unique_daily_coverage * campaign_days)
        
        return base_price, discount, final_price, total_reach, unique_daily_coverage, spots_per_day, total_coverage_percent
        
    except Exception as e:
        logger.error(f"Ошибка расчета стоимости: {e}")
        return 0, 0, MIN_BUDGET, 0, 0, 0, 0

def get_branded_section_name(section):
    names = {
        "auto": "Авторубрики (+20%)",
        "realty": "Недвижимость (+15%)",
        "medical": "Медицинские рубрики (+25%)",
        "custom": "Индивидуальная рубрика (+30%)"
    }
    return names.get(section, "Не выбрана")

def get_time_slots_text(selected_slots):
    """Получить текстовое представление выбранных слотов"""
    slots_text = ""
    for slot_index in selected_slots:
        if 0 <= slot_index < len(TIME_SLOTS_DATA):
            slot = TIME_SLOTS_DATA[slot_index]
            premium_emoji = "🚀" if slot["premium"] else "📊"
            slots_text += f"• {slot['time']} - {slot['label']} {premium_emoji}\n"
    return slots_text

def get_time_slots_detailed_text(selected_slots):
    """Получить детальное представление слотов с охватом"""
    slots_text = ""
    total_coverage = 0
    
    for slot_index in selected_slots:
        if 0 <= slot_index < len(TIME_SLOTS_DATA):
            slot = TIME_SLOTS_DATA[slot_index]
            premium_emoji = "🚀" if slot["premium"] else "📊"
            coverage_percent = slot["coverage_percent"]
            total_coverage += coverage_percent
            slots_text += f"• {slot['time']} - {slot['label']}: {coverage_percent}% {premium_emoji}\n"
    
    return slots_text, total_coverage

def create_excel_file_from_db(campaign_number):
    try:
        logger.info(f"🔍 Начинаем создание Excel для кампании #{campaign_number}")
        
        conn = sqlite3.connect("campaigns.db")
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM campaigns WHERE campaign_number = ?", (campaign_number,))
        campaign_data = cursor.fetchone()
        conn.close()
        
        if not campaign_data:
            logger.error(f"❌ Кампания #{campaign_number} не найдена в БД")
            return None
            
        logger.info(f"✅ Кампания #{campaign_number} найдена в БД")
        
        user_data = {
            "selected_radios": campaign_data[3].split(","),
            "start_date": campaign_data[4],
            "end_date": campaign_data[5],
            "campaign_days": campaign_data[6],
            "selected_time_slots": list(map(int, campaign_data[7].split(","))) if campaign_data[7] else [],
            "branded_section": campaign_data[8],
            "campaign_text": campaign_data[9],
            "production_option": campaign_data[10],
            "contact_name": campaign_data[11],
            "company": campaign_data[12],
            "phone": campaign_data[13],
            "email": campaign_data[14],
            "duration": campaign_data[15],
            "production_cost": PRODUCTION_OPTIONS.get(campaign_data[10], {}).get("price", 0)
        }
        
        logger.info(f"📊 Данные пользователя подготовлены: {len(user_data.get('selected_radios', []))} радиостанций")
        
        base_price, discount, final_price, total_reach, daily_coverage, spots_per_day, total_coverage_percent = calculate_campaign_price_and_reach(user_data)
        logger.info(f"💰 Расчет стоимости: база={base_price}, скидка={discount}, итого={final_price}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Медиаплан {campaign_number}"
        
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=12)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                       top=Side(style="thin"), bottom=Side(style="thin"))
        
        ws.merge_cells("A1:F1")
        ws["A1"] = f"МЕДИАПЛАН КАМПАНИИ #{campaign_number}"
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill
        ws["A1"].alignment = Alignment(horizontal="center")
        
        ws.merge_cells("A2:F2")
        ws["A2"] = "РАДИО ТЮМЕНСКОЙ ОБЛАСТИ"
        ws["A2"].font = Font(bold=True, size=12, color="366092")
        ws["A2"].alignment = Alignment(horizontal="center")
        
        ws.merge_cells("A4:F4")
        ws["A4"] = "✅ Ваша заявка принята! Спасибо за доверие!"
        ws["A4"].font = Font(bold=True, size=11)
        
        ws["A6"] = "📊 ПАРАМЕТРЫ КАМПАНИИ:"
        ws["A6"].font = title_font
        
        # Детальная информация о выбранных слотах
        slots_text, total_coverage = get_time_slots_detailed_text(user_data.get("selected_time_slots", []))
        
        params = [
            f"Радиостанции: {', '.join(user_data.get('selected_radios', []))}",
            f"Период: {user_data.get('start_date')} - {user_data.get('end_date')} ({user_data.get('campaign_days')} дней)",
            f"Выходов в день: {spots_per_day}",
            f"Всего выходов за период: {spots_per_day * user_data.get('campaign_days', 30)}",
            f"Хронометраж ролика: {user_data.get('duration', 20)} сек",
            f"Брендированная рубрика: {get_branded_section_name(user_data.get('branded_section'))}",
            f"Производство: {PRODUCTION_OPTIONS.get(user_data.get('production_option', 'ready'), {}).get('name', 'Не выбрано')}",
            f"Суммарный охват выбранных слотов: {total_coverage}%"
        ]
        
        for i, param in enumerate(params, 7):
            ws[f"A{i}"] = f"• {param}"
        
        ws["A16"] = "📻 ВЫБРАННЫЕ РАДИОСТАНЦИИ:"
        ws["A16"].font = title_font
        
        row = 17
        total_listeners = 0
        for radio in user_data.get("selected_radios", []):
            listeners = STATION_COVERAGE.get(radio, 0)
            total_listeners += listeners
            ws[f"A{row}"] = f"• {radio}: ~{format_number(listeners)} слушателей"
            row += 1
        
        ws[f"A{row}"] = f"• ИТОГО: ~{format_number(total_listeners)} слушателей"
        ws[f"A{row}"].font = Font(bold=True)
        
        row += 2
        ws[f"A{row}"] = "🕒 ВЫБРАННЫЕ ВРЕМЕННЫЕ СЛОТЫ:"
        ws[f"A{row}"].font = title_font
        
        row += 1
        for slot_index in user_data.get("selected_time_slots", []):
            if 0 <= slot_index < len(TIME_SLOTS_DATA):
                slot = TIME_SLOTS_DATA[slot_index]
                ws[f"A{row}"] = f"• {slot['time']} - {slot['label']}: {slot['coverage_percent']}%"
                row += 1
        
        ws[f"A{row}"] = f"• Суммарный охват слотов: {total_coverage}%"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        
        row += 1
        ws[f"A{row}"] = "🎯 РАСЧЕТНЫЙ ОХВАТ:"
        ws[f"A{row}"].font = title_font
        
        row += 1
        ws[f"A{row}"] = f"• Выходов в день: {spots_per_day}"
        row += 1
        ws[f"A{row}"] = f"• Уникальных слушателей в день: ~{format_number(daily_coverage)} чел."
        row += 1
        ws[f"A{row}"] = f"• Общий охват за период: ~{format_number(total_reach)} чел."
        
        row += 2
        ws[f"A{row}"] = "💰 ФИНАНСОВАЯ ИНФОРМАЦИЯ:"
        ws[f"A{row}"].font = title_font
        
        financial_data = [
            ["Позиция", "Сумма (₽)"],
            ["Эфирное время", base_price - user_data.get("production_cost", 0)],
            ["Производство ролика", user_data.get("production_cost", 0)],
            ["", ""],
            ["Базовая стоимость", base_price],
            ["Скидка 50%", -discount],
            ["", ""],
            ["ИТОГО", final_price]
        ]
        
        for i, (item, value) in enumerate(financial_data, row + 1):
            ws[f"A{i}"] = item
            if isinstance(value, int):
                ws[f"B{i}"] = value
                if item == "ИТОГО":
                    ws[f"B{i}"].font = Font(bold=True, color="FF0000")
                elif item == "Скидка 50%":
                    ws[f"B{i}"].font = Font(color="00FF00")
            else:
                ws[f"B{i}"] = value
        
        row = i + 3
        ws[f"A{row}"] = "👤 ВАШИ КОНТАКТЫ:"
        ws[f"A{row}"].font = title_font
        
        contacts = [
            f"Имя: {user_data.get('contact_name', 'Не указано')}",
            f"Телефон: {user_data.get('phone', 'Не указан')}",
            f"Email: {user_data.get('email', 'Не указан')}",
            f"Компания: {user_data.get('company', 'Не указана')}"
        ]
        
        for i, contact in enumerate(contacts, row + 1):
            ws[f"A{i}"] = f"• {contact}"
        
        row = i + 2
        ws[f"A{row}"] = "📞 НАШИ КОНТАКТЫ:"
        ws[f"A{row}"].font = title_font
        ws[f"A{row + 1}"] = "• Email: a.khlistunov@gmail.com"
        ws[f"A{row + 2}"] = "• Telegram: t.me/AlexeyKhlistunov"
        
        row = row + 4
        ws[f"A{row}"] = "🎯 СТАРТ КАМПАНИИ:"
        ws[f"A{row}"].font = title_font
        ws[f"A{row + 1}"] = "В течение 3 рабочих дней после подтверждения"
        
        row = row + 3
        ws[f"A{row}"] = f"📅 Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 15
        
        table_start = row - len(financial_data) - 1
        table_end = table_start + len(financial_data) - 1
        for row_num in range(table_start, table_end + 1):
            for col in ["A", "B"]:
                ws[f"{col}{row_num}"].border = border
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"✅ Excel файл успешно создан для кампании #{campaign_number}, размер: {len(buffer.getvalue())} байт")
        return buffer
        
    except Exception as e:
        logger.error(f"❌ Ошибка при создании Excel: {e}")
        return None

async def send_admin_notification(context, user_data, campaign_number):
    try:
        excel_buffer = create_excel_file_from_db(campaign_number)
        if excel_buffer:
            await context.bot.send_document(
                chat_id=ADMIN_TELEGRAM_ID,
                document=excel_buffer,
                filename=f"mediaplan_{campaign_number}.xlsx",
                caption=f"📊 Медиаплан кампании #{campaign_number}"
            )
            logger.info(f"✅ Excel автоматически отправлен админу для кампании #{campaign_number}")
        
        base_price, discount, final_price, total_reach, daily_coverage, spots_per_day, total_coverage_percent = calculate_campaign_price_and_reach(user_data)
        
        stations_text = ""
        for radio in user_data.get("selected_radios", []):
            listeners = STATION_COVERAGE.get(radio, 0)
            stations_text += f"• {radio}: ~{format_number(listeners)} слушателей\n"
        
        slots_text = get_time_slots_text(user_data.get("selected_time_slots", []))
        
        notification_text = f"""
🔔 НОВАЯ ЗАЯВКА #{campaign_number}

👤 КЛИЕНТ:
Имя: {user_data.get('contact_name', 'Не указано')}
Телефон: {user_data.get('phone', 'Не указан')}
Email: {user_data.get('email', 'Не указан')}
Компания: {user_data.get('company', 'Не указана')}

📊 РАДИОСТАНЦИИ:
{stations_text}
📅 ПЕРИОД: {user_data.get('start_date')} - {user_data.get('end_date')} ({user_data.get('campaign_days')} дней)
🕒 ВЫБРАНО СЛОТОВ: {len(user_data.get('selected_time_slots', []))}
{slots_text}
🎙️ РУБРИКА: {get_branded_section_name(user_data.get('branded_section'))}
⏱️ РОЛИК: {PRODUCTION_OPTIONS.get(user_data.get('production_option', 'ready'), {}).get('name', 'Не выбрано')}
📏 ХРОНОМЕТРАЖ: {user_data.get('duration', 20)} сек

💰 СТОИМОСТЬ:
Базовая: {format_number(base_price)}₽
Скидка 50%: -{format_number(discount)}₽
Итоговая: {format_number(final_price)}₽

🎯 РАСЧЕТНЫЙ ОХВАТ:
• Выходов в день: {spots_per_day}
• Всего выходов: {spots_per_day * user_data.get('campaign_days', 30)}
• Ежедневно: ~{format_number(daily_coverage)} чел.
• За период: ~{format_number(total_reach)} чел.
"""
        
        keyboard = [
            [
                InlineKeyboardButton(f"📞 {user_data.get('phone', 'Телефон')}", callback_data=f"call_{user_data.get('phone', '')}"),
                InlineKeyboardButton(f"✉️ Написать", callback_data=f"email_{user_data.get('email', '')}")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await context.bot.send_message(
            chat_id=ADMIN_TELEGRAM_ID,
            text=notification_text,
            reply_markup=reply_markup
        )
        logger.info(f"✅ Уведомление админу отправлено для кампании #{campaign_number}")
        return True
        
    except Exception as e:
        logger.error(f"❌ Ошибка отправки админу: {e}")
        return False
